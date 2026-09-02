"""
tidy_excel_operasional.py - rapihkan 15 sheet OPERASIONAL di
"Monitoring EX PRODUKSI 2026.xlsx" yang TIDAK punya padanan di skema web
(surat jalan subcon, upah borongan, PO aksesoris, rekap pengiriman, dll.
-> memetakan ke modul legacy spk_produksi yang dorman).

Karena tak ada tabel web untuk dibandingkan, ini murni "rapihkan":
  - pindah header ke baris 1 (header asli biasanya di baris 5/6/12 setelah kop surat)
  - betulkan typo header (So Oder->so_order, Itim->item, Finsh->finishing,
    Permarks->remarks, tiga kolom "-/+" diberi nama unik, Nominal AKhir->nominal_akhir,
    cat han->catatan) + slug ke snake_case, kolom kosong -> col_2/col_3...
  - buang baris TOTAL/subtotal per-blok, baris duplikat-header, padding ribuan baris kosong
  - strip spasi tersembunyi di nilai teks ("Yard " -> "Yard"); sentinel #REF!/#VALUE!/NULL/- -> kosong
  - bulatkan noise presisi float (4 dp)
  - Gudang Dewi: buang kolom yang >50% selnya #REF! (kolom size XS-6XL & aksesoris
    semuanya =#REF!); catat bahwa datanya tak bisa dipulihkan dari file ini
  - Sheet28 (kosong): dilewati, dicatat "hapus"

Output:
  <downloads>/Monitoring EX PRODUKSI 2026 - RAPI (operasional).xlsx   (1 sheet bersih per sheet sumber)
  backend/database/tidy_blueprint_2026/RAPI_OPERASIONAL.md            (log aksi per sheet)

Pakai:
  cd backend
  python scripts/tidy_excel_operasional.py \
    "C:/Users/borde/Downloads/skrip/Salinan dari Monitoring EX PRODUKSI 2026.xlsx"
"""
from __future__ import annotations

import re
import sys
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from scripts.import_excel_blueprint import num, s  # noqa: E402

import openpyxl  # noqa: E402
from openpyxl.styles import Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

HERE = Path(__file__).resolve().parent
OUTDIR = HERE.parent / "database" / "tidy_blueprint_2026"
DOWNLOADS = Path.home() / "Downloads" / "skrip"

# sheet -> baris header (1-based) hasil analisa struktur. None = auto-deteksi.
HEADER_ROW = {
    "Gudang Dewi": 5, "FORM WI": 6, "Pengajaun Akse": 12, "finishing": 4,
    "Output Cutt Pers": 5, "Form Produksi Anis": 5, "To Produksi Pa Ato": 5,
    "Rekapan Pengiriman Sandi": 5, "To Washing Anis": 5,
    "To Print Mentah Puring Pa Ato": 5, "To Embrodeiry  Mentah Pa Ato": 5,
    "To Embrodeiry Jadi  Anis": 5, "Surat Jalan Sample": None,
    "Finishing": 5, "Finishing Borongan _": 5,
}
# nama sheet keluaran yang lebih jelas (sumber punya "finishing" & "Finishing " sekaligus)
OUT_NAME = {"finishing": "finishing (3 tabel)", "Finishing": "Finishing (setoran)"}
# sheet yang PUNYA padanan web -> ditangani tidy_excel_blueprint.py, dilewati di sini
WEB_SHEETS = {"Monitoring", "GUDANG BAHAN", "FORM RIJEK BORDIR", "DATA"}
EMPTY_SHEETS = {"Sheet28"}

HEADER_FIX = {
    "so oder": "so_order", "so order": "so_order", "so": "so", "so odER": "so_order",
    "itim": "item", "finsh": "finishing", "permarks": "remarks", "remark": "remarks",
    "nominal akhir": "nominal_akhir", "nominal akhir ": "nominal_akhir",
    "cat han": "catatan", "no surat jalan": "no_surat_jalan", "tgl update": "tgl_update",
    "hasil prod": "hasil_prod", "total rijek": "total_rijek",
}


def slug(v, i, seen):
    t = " ".join(str(v or "").split()).strip().lower()
    t = HEADER_FIX.get(t, t)
    if not t or t in {"-", ".", "none"}:
        t = f"col_{i + 1}"
    t = re.sub(r"[^\w]+", "_", t, flags=re.U).strip("_").lower() or f"col_{i + 1}"
    base, k = t, t
    n = 2
    while k in seen:
        k = f"{base}_{n}"
        n += 1
    seen.add(k)
    return k


def clean_val(v):
    if isinstance(v, (datetime, date)):
        return v.isoformat()[:10]
    if isinstance(v, float):
        return round(v, 4)
    t = s(v)                       # trim + sentinel #REF!/#VALUE!/NULL/- -> ""
    if t == "":
        return None
    n = num(t)
    return n if (n is not None and re.fullmatch(r"-?[\d.,]+", t)) else t


def is_total_row(cells):
    joined = " ".join(str(c) for c in cells[:3] if c).upper()
    if re.search(r"\bTOTAL\b|\bJUMLAH\b|\bGRAND\b|\bSUB ?TOTAL\b", joined):
        nums = sum(1 for c in cells if isinstance(c, (int, float)))
        return nums >= 1
    return False


def auto_header(rows):
    best, score = 0, -1
    for i, r in enumerate(rows[:14]):
        cells = [c for c in r if c is not None and str(c).strip()]
        txt = [c for c in cells if isinstance(c, str) and not re.fullmatch(r"-?[\d.,]+", c.strip())]
        sc = len(txt) - (0 if i >= 3 else 1)
        if len(cells) >= 4 and sc > score:
            best, score = i, sc
    return best


def tidy_sheet(ws, log):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        log.append(f"- **{ws.title}**: kosong, dilewati.")
        return None, None
    hr = HEADER_ROW.get(ws.title)
    hidx = (hr - 1) if hr else auto_header(rows)
    raw_header = rows[hidx] if hidx < len(rows) else rows[0]
    ncol = max((i + 1 for i, c in enumerate(raw_header) if c is not None and str(c).strip()),
               default=len(raw_header))
    seen = set()
    header = [slug(raw_header[i] if i < len(raw_header) else None, i, seen) for i in range(ncol)]

    out, dup_hdr, totals, blanks, frag = [], 0, 0, 0, 0
    hdr_norm = [h for h in header if not h.startswith("col_")]
    # kolom kunci utk buang fragmen kop-surat di sheet berbentuk log-pengiriman
    key_i = next((i for i, h in enumerate(header) if h in ("so_order", "so")), None)
    date_i = next((i for i, h in enumerate(header) if h in ("tanggal", "tgl_update")), None)
    def empty(c):
        return c is None or str(c).strip() in ("", "0", "0.0")

    for r in rows[hidx + 1:]:
        cells = [clean_val(r[i]) if i < len(r) else None for i in range(ncol)]
        if all(empty(c) for c in cells):        # baris kosong / template VLOOKUP kosong
            blanks += 1
            continue
        low = [str(c).strip().lower() for c in cells if c is not None]
        if sum(1 for h in hdr_norm if h.replace("_", " ") in low) >= max(3, len(hdr_norm) // 2):
            dup_hdr += 1
            continue
        if is_total_row(cells):
            totals += 1
            continue
        # fragmen kop-surat: tak ada SO & tak ada tanggal & isi tipis
        if key_i is not None and empty(cells[key_i]) and (date_i is None or empty(cells[date_i])) \
                and sum(1 for c in cells if not empty(c)) < 4:
            frag += 1
            continue
        out.append(cells)

    # buang kolom yang mayoritas #REF! (Gudang Dewi) -> clean_val sudah -> None,
    # jadi deteksi kolom yang ~seluruhnya None padahal header-nya "berisi"
    drop = []
    if out:
        for ci in range(ncol):
            nonnull = sum(1 for row in out if row[ci] is not None)
            if nonnull == 0:          # kolom 100% kosong (#REF!/spacer/VLOOKUP mati)
                drop.append(ci)
    coldrop = ""
    if drop:
        keep = [i for i in range(ncol) if i not in drop]
        header = [header[i] for i in keep]
        out = [[row[i] for i in keep] for row in out]
        coldrop = f", {len(drop)} kolom kosong/#REF! ({[str(raw_header[i]) for i in drop][:6]})"
    log.append(f"- **{ws.title}**: header baris {hidx + 1}; {len(out)} baris bersih "
               f"(buang {blanks} kosong, {totals} TOTAL, {dup_hdr} ulang-header, "
               f"{frag} fragmen kop-surat{coldrop}).")
    return header, out


HEAD_FILL = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(bold=True, color="FFFFFF")


def main():
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else \
        DOWNLOADS / "Salinan dari Monitoring EX PRODUKSI 2026.xlsx"
    if not src.exists():
        sys.exit(f"ERROR: tidak ketemu: {src}")

    wb_in = openpyxl.load_workbook(src, data_only=True, read_only=True)
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)
    log = [f"# Rapihkan sheet operasional - {src.name}", "",
           f"Dibuat: {datetime.now().isoformat(timespec='seconds')}  ", "",
           "Sheet ini **tidak** punya padanan di skema web (mengarah ke modul legacy "
           "`spk_produksi` yang dorman), jadi hanya dirapihkan - tidak direkonsiliasi "
           "dengan prod. Sheet yang punya padanan web (`Monitoring`, `GUDANG BAHAN`, "
           "`FORM RIJEK BORDIR`, `DATA`) ditangani `tidy_excel_blueprint.py`.", "",
           "## Aksi per sheet", ""]
    done = 0
    for name in wb_in.sheetnames:
        if name in WEB_SHEETS:
            log.append(f"- **{name}**: punya padanan web -> lihat workbook RAPI utama.")
            continue
        if name in EMPTY_SHEETS:
            log.append(f"- **{name}**: sheet kosong -> aman dihapus.")
            continue
        ws = wb_in[name]
        header, rows = tidy_sheet(ws, log)
        if not header:
            continue
        title = OUT_NAME.get(name.strip(), re.sub(r"\s+", " ", name).strip())[:31]
        wo = wb_out.create_sheet(title)
        wo.append(header)
        for c in range(1, len(header) + 1):
            wo.cell(row=1, column=c).fill = HEAD_FILL
            wo.cell(row=1, column=c).font = HEAD_FONT
        for row in rows:
            wo.append(row)
        wo.freeze_panes = "A2"
        for i, h in enumerate(header, 1):
            w = max([len(str(h))] + [len(str(r[i - 1])) for r in rows[:200] if r[i - 1] is not None] or [0])
            wo.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 55)
        done += 1
    wb_in.close()

    info = wb_out.create_sheet("_INFO", 0)
    info.append(["sheet", "baris bersih", "kolom"])
    for c in range(1, 4):
        info.cell(row=1, column=c).fill = HEAD_FILL
        info.cell(row=1, column=c).font = HEAD_FONT
    for wsx in wb_out.worksheets:
        if wsx.title != "_INFO":
            info.append([wsx.title, wsx.max_row - 1, wsx.max_column])
    info.column_dimensions["A"].width = 30

    out_xlsx = DOWNLOADS / "Monitoring EX PRODUKSI 2026 - RAPI (operasional).xlsx"
    wb_out.save(out_xlsx)
    log += ["", f"## Hasil", "",
            f"- Workbook: `{out_xlsx.name}` ({done} sheet bersih + _INFO)",
            "- Sheet form (kop surat + blok berulang: `Pengajaun Akse`, `Surat Jalan Sample`, "
            "`FORM WI`) tetap ber-bentuk form; baris item diekstrak apa adanya di bawah "
            "header yang terdeteksi - perlu rapi manual bila mau jadi tabel relasional.",
            "- `Gudang Dewi`: kolom ukuran (XS-6XL) & aksesoris hilang permanen (semua `=#REF!` "
            "di file sumber, turunan dari sheet lain yang barisnya sudah bergeser). Kolom yang "
            "selamat (No, Tanggal, So Order, Brand, Artikel, Style, Maklon, blok kanan "
            "Name Bahan/Yard/Harga/Nominal) dipertahankan.",
            "- File 1 (`DATA BAHAN NEW 2026.xlsx`) juga punya 3 sheet non-web (`REKAPAN` rekap, "
            "`FABRIC INSPECTION` form, `Sheet25` surat jalan) - kecil, belum dirapihkan di ronde ini."]
    (OUTDIR / "RAPI_OPERASIONAL.md").write_text("\n".join(log), encoding="utf-8")

    print(f"OK -> {out_xlsx}  ({out_xlsx.stat().st_size:,} B), {done} sheet")
    print(f"log -> {OUTDIR / 'RAPI_OPERASIONAL.md'}")


if __name__ == "__main__":
    main()
