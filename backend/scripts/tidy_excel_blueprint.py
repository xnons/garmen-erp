"""
tidy_excel_blueprint.py - hasilkan workbook Excel RAPI + rekonsiliasi dengan DB web (prod).

Memakai ULANG parser bersih dari `import_excel_blueprint.py` (parse_bahan /
parse_monitoring) - semua aturan pembersihan (sentinel #REF!/#VALUE!/NULL/-,
dedup kunci alami, koersi angka & tanggal, kanonikalisasi nama buyer/subcon,
stok negatif -> 0, gerbang baris sampah, baris TOTAL diabaikan) sudah ada di sana.

Skrip ini menambax:
  - buang kolom helper (_brand/_status/_tanggal/_color/_warna/_jenis_bahan)
  - namai & urutkan kolom mengikuti tampilan web
  - banding tiap baris dengan snapshot prod (JSON read-only di _prod_snapshot_2026-09-02/)
  - kolom `_REKONSILIASI` per baris
  - aturan "EXCEL MENANG": nilai Excel yang ditulis; tiap selisih -> perubahan_prod.sql
    (idempoten, natural key, TIDAK dijalankan otomatis).
    Pengecualian sengaja: sel UANG kosong/0 di Excel != "Excel bilang 0" -
    kalau prod terisi, pertahankan angka prod & tandai PERLU KEPUTUSAN, tanpa UPDATE.
  - baris yang hanya ada di prod -> sheet "HANYA DI WEB" (tidak pernah DELETE)

Output:
  <downloads>/DATA BAHAN NEW 2026 - RAPI.xlsx
  <downloads>/Monitoring EX PRODUKSI 2026 - RAPI.xlsx
  backend/database/tidy_blueprint_2026/*.csv           (+ kolom _rekonsiliasi)
  backend/database/tidy_blueprint_2026/REKONSILIASI_BAHAN.md
  backend/database/tidy_blueprint_2026/REKONSILIASI_PRODUKSI.md
  backend/database/tidy_blueprint_2026/perubahan_prod.sql

Pakai:
  cd backend
  python scripts/tidy_excel_blueprint.py \
    "C:/Users/borde/Downloads/skrip/Salinan dari DATA BAHAN NEW 2026.xlsx" \
    "C:/Users/borde/Downloads/skrip/Salinan dari Monitoring EX PRODUKSI 2026.xlsx"
  (tanpa argumen -> pakai path default di ~/Downloads/skrip/)
"""
from __future__ import annotations

import csv
import json
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

# --- import parser + helper dari importer (jangan salin logikanya) --------------
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from scripts.import_excel_blueprint import (  # noqa: E402
    canon_buyer, parse_bahan, parse_monitoring,
)

import openpyxl  # noqa: E402
from openpyxl.styles import Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

HERE = Path(__file__).resolve().parent
OUTDIR = HERE.parent / "database" / "tidy_blueprint_2026"
SNAPDIR = OUTDIR / "_prod_snapshot_2026-09-02"
DOWNLOADS = Path.home() / "Downloads" / "skrip"

# --------------------------------------------------------------------------- #
# Kolom rapi per entitas (urut, mengikuti web / CSV importer)
# --------------------------------------------------------------------------- #
COLS = {
    "inventory_items": ["item_code", "description", "item_type", "unit", "unit_price",
                        "current_stock", "min_stock_alert", "width_inch", "rack_location"],
    "material_receipts": ["item_code", "receipt_date", "roll_number", "qty_received",
                          "unit", "contract_type", "inspection_status"],
    "material_allocations": ["so_number", "item_code", "dispatch_date", "qty_issued",
                             "surat_jalan_no"],
    "sales_orders_codeso": ["so_number", "brand", "style_name"],
    "sales_orders": ["so_number", "brand", "style_name", "item_category", "color",
                     "order_qty", "status", "order_date", "special_instructions"],
    "wip_movements": ["so_number", "stage_name", "sequence_order", "vendor",
                      "surat_jalan_no", "dispatch_date", "qty_dispatched",
                      "received_date", "qty_received", "qty_reject",
                      "balance_discrepancy", "status"],
    "cutting_records": ["so_number", "cutting_date", "qty_cut", "main_fabric_used",
                        "puring_used", "main_consumption_rate", "puring_consumption_rate"],
    "reject_logs": ["so_number", "stage_name", "defect_reason", "qty_reject",
                    "unit_cost_loss", "total_loss"],
    "partners": ["name", "category"],
}
MONEY_COLS = {"unit_price", "unit_cost_loss", "total_loss"}
DATE_COLS = {"receipt_date", "dispatch_date", "order_date", "cutting_date", "received_date"}

# Kolom NUMERIK di mana "kosong / 0 di Excel" TIDAK berarti "Excel bilang 0".
# Kalau Excel 0 tapi prod terisi -> pertahankan angka prod, tandai PERLU KEPUTUSAN,
# JANGAN emit UPDATE yang menol-kan. (Selaras perilaku importer utk unit_price.)
KEEP_PROD_IF_ZERO = {
    "inventory_items": {"unit_price"},
    "cutting_records": {"main_fabric_used", "puring_used",
                        "main_consumption_rate", "puring_consumption_rate"},
    "reject_logs": {"unit_cost_loss", "total_loss"},
}


def num_delta(field, ex, pr, entity):
    """(mode, note) - mode: 'same' | 'excel' (Excel menang) | 'keepprod' (pertahankan prod)."""
    exf, prf = float(ex or 0), float(pr or 0)
    if approx(exf, prf):
        return "same", None
    if exf == 0 and prf > 0 and field in KEEP_PROD_IF_ZERO.get(entity, ()):
        return "keepprod", f"PERLU KEPUTUSAN: {field} kosong di Excel, prod={prf} (dipertahankan)"
    return "excel", f"BEDA {field}: excel={exf} prod={prf}"


# --------------------------------------------------------------------------- #
# Util
# --------------------------------------------------------------------------- #
def norm(v) -> str:
    return " ".join(str(v or "").split()).upper()


def approx(a, b, tol=0.01) -> bool:
    try:
        return abs(float(a or 0) - float(b or 0)) <= tol
    except (TypeError, ValueError):
        return norm(a) == norm(b)


def isod(v) -> str:
    if isinstance(v, (date, datetime)):
        return v.isoformat()[:10]
    return str(v or "")[:10]


def sqlstr(v) -> str:
    if v is None or v == "":
        return "NULL"
    return "'" + str(v).replace("'", "''") + "'"


def clean_row(d: dict, entity: str) -> dict:
    """Ambil hanya kolom rapi, buang helper _*."""
    cols = COLS[entity]
    return {c: d.get(c) for c in cols}


def load_snap(name: str):
    return json.loads((SNAPDIR / name).read_text(encoding="utf-8"))


# --------------------------------------------------------------------------- #
# Pengumpul perubahan prod + laporan
# --------------------------------------------------------------------------- #
class Changes:
    def __init__(self):
        self.inv, self.cut, self.so_safe, self.so_risky, self.ins = [], [], [], [], []
        self.notes = []

    def dump(self, path: Path):
        L = ["-- " + "=" * 70,
             "-- PERUBAHAN UNTUK PROD  -  REVIEW MANUAL. JANGAN dijalankan otomatis.",
             f"-- Dibuat: {datetime.now().isoformat(timespec='seconds')} oleh tidy_excel_blueprint.py",
             "-- Aturan: \"Excel menang\" - tiap UPDATE mengembalikan nilai prod ke angka",
             "-- Excel yang sudah dibersihkan. Kunci natural (item_code / so_number).",
             "-- Idempoten: aman dijalankan ulang. TIDAK ada DELETE.",
             "-- " + "=" * 70, "", "BEGIN;", ""]
        L.append(f"-- A. inventory_items - stok / harga / deskripsi  ({len(self.inv)} baris)")
        L += self.inv or ["--   (tidak ada perubahan)"]
        L.append("")
        L.append(f"-- B. cutting_records - qty_cut / konsumsi kain  ({len(self.cut)} baris)")
        L += self.cut or ["--   (tidak ada perubahan)"]
        L.append("")
        L.append(f"-- C. sales_orders - field aman: style_name / item_category / color / brand"
                 f"  ({len(self.so_safe)} baris)")
        L += self.so_safe or ["--   (tidak ada perubahan)"]
        L.append("")
        L.append("-- D. sales_orders - KONSEKUENSIAL: buyer / status / order_qty dari sheet.")
        L.append("--    UPDATE di bawah bisa mengganti buyer atau MENIMPA progres yang mungkin")
        L.append("--    sah dari pemakaian aplikasi. buyer_id pakai COALESCE (tak pernah jadi")
        L.append(f"--    NULL). Tinjau baris per baris sebelum jalan.  ({len(self.so_risky)} baris)")
        L += self.so_risky or ["--   (tidak ada perubahan)"]
        L.append("")
        L.append(f"-- E. Baris BARU (ada di Excel bersih, belum di prod)  ({len(self.ins)} baris)")
        L += self.ins or ["--   (tidak ada)"]
        L.append("")
        for n in self.notes:
            L.append(f"-- {n}")
        L.append("")
        L.append("COMMIT;")
        L.append("")
        path.write_text("\n".join(L), encoding="utf-8")


# --------------------------------------------------------------------------- #
# REKONSILIASI per entitas  -> (rows_with_note, summary_dict, hanya_di_web_list)
# --------------------------------------------------------------------------- #
def rec_inventory(items, ch: Changes):
    prod = {p["item_code"]: p for p in load_snap("inventory_items.json")}
    rows, web_only = [], []
    ok = beda = baru = keputusan = 0
    for it in items:
        r = clean_row(it, "inventory_items")
        code = r["item_code"]
        p = prod.get(code)
        notes = []
        if not p:
            baru += 1
            notes.append("BARU: item belum ada di prod")
            ch.ins.append(
                f"INSERT INTO inventory_items (id,item_code,description,item_type,unit,"
                f"unit_price,current_stock,min_stock_alert,width_inch,gramasi_gsm,rack_location) "
                f"SELECT gen_random_uuid()::text,{sqlstr(code)},{sqlstr(r['description'])},"
                f"{sqlstr(r['item_type'])},{sqlstr(r['unit'])},{float(r['unit_price'] or 0)},"
                f"{float(r['current_stock'] or 0)},50,58,0,{sqlstr(r['rack_location'])} "
                f"WHERE NOT EXISTS (SELECT 1 FROM inventory_items WHERE item_code={sqlstr(code)});")
        else:
            upd = {}
            if norm(r["description"]) != norm(p["description"]):
                notes.append(f"BEDA description: excel={r['description']!r} prod={p['description']!r}")
                upd["description"] = r["description"]
            if norm(r["item_type"]) != norm(p["item_type"]):
                notes.append(f"BEDA item_type: excel={r['item_type']} prod={p['item_type']}")
                upd["item_type"] = r["item_type"]
            if not approx(r["current_stock"], p["current_stock"]):
                notes.append(f"BEDA current_stock: excel={r['current_stock']} prod={p['current_stock']}")
                upd["current_stock"] = float(r["current_stock"] or 0)
            # unit_price - pengecualian sel uang kosong
            ep, pp = float(r["unit_price"] or 0), float(p["unit_price"] or 0)
            if ep > 0 and not approx(ep, pp):
                notes.append(f"BEDA unit_price: excel={ep} prod={pp}")
                upd["unit_price"] = ep
            elif ep == 0 and pp > 0:
                keputusan += 1
                notes.append(f"PERLU KEPUTUSAN: harga kosong di Excel, prod={pp} (dipertahankan)")
                r["unit_price"] = pp  # tampilkan angka prod di sheet rapi
            if upd:
                beda += 1
                sets = ", ".join(
                    f"{k}={sqlstr(v) if k in ('description','item_type') else v}"
                    for k, v in upd.items())
                ch.inv.append(f"UPDATE inventory_items SET {sets} WHERE item_code={sqlstr(code)};")
            else:
                ok += 1
        r["_REKONSILIASI"] = " | ".join(notes) if notes else "OK - sama dengan prod"
        rows.append(r)
    excel_codes = {it["item_code"] for it in items}
    for code, p in sorted(prod.items()):
        if code not in excel_codes:
            web_only.append(("inventory_items", code, json.dumps(p, ensure_ascii=False)))
    return rows, {"excel": len(items), "OK": ok, "BEDA": beda, "BARU": baru,
                  "PERLU_KEPUTUSAN": keputusan, "HANYA_DI_WEB": len(web_only)}, web_only


def rec_receipts(receipts, ch: Changes):
    fab = load_snap("fabric.json")
    imp = load_snap("imp_maps.json")["receipts_imp"]  # {roll_number: qty}
    rows, ok, beda, baru = [], 0, 0, 0
    for rec in receipts:
        r = clean_row(rec, "material_receipts")
        r["receipt_date"] = isod(r["receipt_date"])
        pk = r["roll_number"]
        if pk in imp:
            if approx(r["qty_received"], imp[pk]):
                ok += 1
                note = "OK - baris IMP-2026 di prod = Excel bersih"
            else:
                beda += 1
                note = f"BEDA qty_received: excel={r['qty_received']} prod={imp[pk]}"
        else:
            baru += 1
            note = "BARU: baris IMP-2026 belum ada di prod"
        r["_REKONSILIASI"] = note
        rows.append(r)
    web_only = [("material_receipts", x.get("roll_number") or "(NULL)",
                 json.dumps(x, ensure_ascii=False)) for x in fab["material_receipts_nonimp"]]
    if beda or baru:
        ch.notes.append(f"material_receipts: {beda} beda + {baru} baru vs prod - "
                        f"jalankan ulang bagian 4 dari import_blueprint_2026/import_blueprint.sql")
    else:
        ch.notes.append(f"material_receipts: {ok} baris IMP-2026 di prod sudah = Excel bersih "
                        f"(tidak ada perubahan).")
    return rows, {"excel": len(receipts), "OK": ok, "BEDA": beda, "BARU": baru,
                  "HANYA_DI_WEB": len(web_only)}, web_only


def rec_allocs(allocs, ch: Changes):
    fab = load_snap("fabric.json")
    imp = load_snap("imp_maps.json")["allocs_imp"]  # {surat_jalan_no: qty} (kunci non-unik di prod)
    ex_sum, ex_cnt = defaultdict(float), Counter()
    for a in allocs:
        k = f"IMP-2026-{a['so_number']}-{a['item_code']}"
        ex_sum[k] += float(a["qty_issued"] or 0)
        ex_cnt[k] += 1
    rows, ok, beda, baru = [], 0, 0, 0
    for a in allocs:
        r = clean_row(a, "material_allocations")
        r["dispatch_date"] = isod(r["dispatch_date"])
        k = r["surat_jalan_no"]
        if k in imp:
            if approx(ex_sum[k], imp[k]):
                ok += 1
                note = "OK - baris IMP-2026 di prod = Excel bersih"
                if ex_cnt[k] > 1:
                    note += f" (catatan: {ex_cnt[k]} baris berbagi surat_jalan_no ini)"
            else:
                beda += 1
                note = f"BEDA agregat qty: excel={round(ex_sum[k],2)} prod={imp[k]}"
        else:
            baru += 1
            note = "BARU: baris IMP-2026 belum ada di prod"
        r["_REKONSILIASI"] = note
        rows.append(r)
    web_only = [("material_allocations", x.get("surat_jalan_no") or "(NULL)",
                 json.dumps(x, ensure_ascii=False)) for x in fab["material_allocations_nonimp"]]
    if beda or baru:
        ch.notes.append(f"material_allocations: {beda} beda + {baru} baru vs prod - "
                        f"jalankan ulang bagian 5 dari import_blueprint_2026/import_blueprint.sql")
    else:
        ch.notes.append(f"material_allocations: {len(allocs)} baris ({len(ex_sum)} kunci) di prod "
                        f"sudah = Excel bersih (tidak ada perubahan).")
    return rows, {"excel": len(allocs), "OK": ok, "BEDA": beda, "BARU": baru,
                  "HANYA_DI_WEB": len(web_only)}, web_only


def _prod_so_map():
    return {s["so_number"]: s for s in load_snap("fabric.json")["sales_orders"]}


def rec_codeso(codeso, ch: Changes, mon_numbers):
    """Sheet 'Code So' = prefill minimal. Untuk SO yang juga ada di sheet Monitoring,
    biarkan rekonsiliasi Monitoring yang menangani style_name/brand (importer pun
    menimpa Code So dgn Monitoring)."""
    prod = _prod_so_map()
    rows, ok, beda, baru = [], 0, 0, 0
    for c in codeso:
        r = clean_row(c, "sales_orders_codeso")
        so = r["so_number"]
        p = prod.get(so)
        in_mon = so in mon_numbers
        notes = []
        if not p:
            baru += 1
            notes.append("BARU: SO belum ada di prod")
            if not in_mon:
                ch.ins.append(
                    f"INSERT INTO sales_orders (id,so_number,buyer_id,style_name,item_category,"
                    f"order_qty,status,order_date,contract_type) SELECT gen_random_uuid()::text,"
                    f"{sqlstr(so)},(SELECT id FROM partners WHERE name={sqlstr(canon_buyer(r['brand']))} "
                    f"ORDER BY created_at LIMIT 1),{sqlstr((r['style_name'] or so)[:150])},'GARMENT',0,"
                    f"'REGISTERED','2026-01-01','CMT' WHERE NOT EXISTS "
                    f"(SELECT 1 FROM sales_orders WHERE so_number={sqlstr(so)});")
        else:
            diff = []
            if r["style_name"] and norm(r["style_name"]) != norm(p["style_name"]):
                diff.append(f"BEDA style_name: excel={r['style_name']!r} prod={p['style_name']!r}")
            cb = canon_buyer(r["brand"])
            if cb and norm(cb) != norm(p.get("buyer_name")):
                diff.append(f"BEDA brand: excel={cb} prod_buyer={p.get('buyer_name')}")
            if diff and in_mon:
                notes.append("selisih diselesaikan oleh sheet Monitoring: " + " ; ".join(diff))
            elif diff:
                notes += diff
                if r["style_name"]:
                    ch.so_safe.append(f"UPDATE sales_orders SET style_name={sqlstr(r['style_name'][:150])} "
                                      f"WHERE so_number={sqlstr(so)};  -- prod: {p['style_name']!r}")
                if cb and norm(cb) != norm(p.get("buyer_name")):
                    ch.so_risky.append(
                        f"UPDATE sales_orders SET buyer_id=COALESCE((SELECT id FROM partners WHERE "
                        f"name={sqlstr(cb)} ORDER BY created_at LIMIT 1), buyer_id) "
                        f"WHERE so_number={sqlstr(so)};  -- prod buyer: {p.get('buyer_name')}")
            beda += 1 if notes and not notes[0].startswith("selisih diselesaikan") else 0
            ok += 1 if not notes or notes[0].startswith("selisih diselesaikan") else 0
        r["_REKONSILIASI"] = " | ".join(notes) if notes else "OK - SO ada di prod"
        rows.append(r)
    return rows, {"excel": len(codeso), "OK": ok, "BEDA": beda, "BARU": baru}, []


def rec_sales_orders(sos, ch: Changes, codeso_numbers):
    prod = _prod_so_map()
    rows, ok, beda, baru = [], 0, 0, 0
    for o in sos:
        r = clean_row(o, "sales_orders")
        for k in ("order_date",):
            r[k] = isod(r[k])
        so = r["so_number"]
        p = prod.get(so)
        notes = []
        if not p:
            baru += 1
            notes.append("BARU: SO belum ada di prod")
            ch.ins.append(f"-- SO BARU {so}: buat lewat import_blueprint.sql (bagian 3) atau UI PPIC")
        else:
            safe, risky = [], []
            if r["style_name"] and norm(r["style_name"]) != norm(p["style_name"]):
                notes.append(f"BEDA style_name: excel={r['style_name']!r} prod={p['style_name']!r}")
                safe.append(f"style_name={sqlstr(r['style_name'][:150])}")
            if r["item_category"] and norm(r["item_category"]) != norm(p["item_category"]):
                notes.append(f"BEDA item_category: excel={r['item_category']} prod={p['item_category']}")
                safe.append(f"item_category={sqlstr(r['item_category'])}")
            if r["color"] and norm(r["color"]) != norm(p["color"]):
                notes.append(f"BEDA color: excel={r['color']} prod={p['color']}")
                safe.append(f"color={sqlstr(r['color'])}")
            cb = canon_buyer(o.get("brand"))
            if cb and norm(cb) != norm(p.get("buyer_name")):
                notes.append(f"BEDA brand: excel={cb} prod_buyer={p.get('buyer_name')}")
                ch.so_risky.append(
                    f"UPDATE sales_orders SET buyer_id=COALESCE((SELECT id FROM partners WHERE "
                    f"name={sqlstr(cb)} ORDER BY created_at LIMIT 1), buyer_id) "
                    f"WHERE so_number={sqlstr(so)};  -- prod buyer: {p.get('buyer_name')}")
            if not approx(r["order_qty"], p["order_qty"]):
                notes.append(f"BEDA order_qty: excel={r['order_qty']} prod={p['order_qty']}")
                risky.append(f"order_qty={int(r['order_qty'] or 0)}")
            if norm(r["status"]) != norm(p["status"]):
                notes.append(f"BEDA status: excel={r['status']} prod={p['status']}")
                risky.append(f"status={sqlstr(r['status'])}")
            if safe:
                ch.so_safe.append(f"UPDATE sales_orders SET {', '.join(safe)} WHERE so_number={sqlstr(so)};")
            if risky:
                ch.so_risky.append(f"UPDATE sales_orders SET {', '.join(risky)} "
                                   f"WHERE so_number={sqlstr(so)};  -- prod: "
                                   f"status={p['status']} order_qty={p['order_qty']}")
            beda += 1 if notes else 0
            ok += 0 if notes else 1
        r["_REKONSILIASI"] = " | ".join(notes) if notes else "OK - sama dengan prod"
        rows.append(r)
    excel_so = {o["so_number"] for o in sos} | set(codeso_numbers)
    web_only = [("sales_orders", so, json.dumps(p, ensure_ascii=False))
                for so, p in sorted(prod.items()) if so not in excel_so]
    return rows, {"excel": len(sos), "OK": ok, "BEDA": beda, "BARU": baru,
                  "HANYA_DI_WEB": len(web_only)}, web_only


def rec_wip(wips, ch: Changes):
    prod = load_snap("production.json")
    imp = load_snap("imp_maps.json")["wip_imp"]  # {sjn: [disp,recv,rej,bal,status]}
    rows, ok, beda, baru, flag = [], 0, 0, 0, 0
    for w in wips:
        r = clean_row(w, "wip_movements")
        r["dispatch_date"] = isod(r["dispatch_date"])
        r["received_date"] = isod(r["received_date"])
        k = r["surat_jalan_no"]
        pv = imp.get(k)
        notes = []
        if not pv:
            baru += 1
            notes.append("BARU: baris IMP-2026 belum ada di prod")
        else:
            d, rc, rj, bl, st = pv
            if not approx(r["qty_dispatched"], d):
                notes.append(f"BEDA qty_dispatched: excel={r['qty_dispatched']} prod={d}")
            if not approx(r["qty_received"], rc):
                notes.append(f"BEDA qty_received: excel={r['qty_received']} prod={rc}")
            if not approx(r["qty_reject"], rj):
                notes.append(f"BEDA qty_reject: excel={r['qty_reject']} prod={rj}")
            if norm(r["status"]) != norm(st):
                notes.append(f"BEDA status: excel={r['status']} prod={st}")
            if notes:
                beda += 1
            else:
                ok += 1
        if norm(r["status"]) == "DISCREPANCY_FLAG":
            flag += 1
            notes.append("DISCREPANCY_FLAG - setia pada angka sumber (bukan bug impor)")
        r["_REKONSILIASI"] = " | ".join(notes) if notes else "OK - sama dengan prod"
        rows.append(r)
    imp_sjn = {x["surat_jalan_no"] for x in prod["wip_movements"] if x["is_imp"]}
    web_only = [("wip_movements", x["surat_jalan_no"] or f"{x['so_number']}/{x['stage_name']}",
                 json.dumps({kk: x[kk] for kk in ("so_number", "stage_name", "qty_dispatched",
                            "qty_received", "qty_reject", "status", "partner_name")},
                           ensure_ascii=False))
                for x in prod["wip_movements"] if not x["is_imp"]]
    if beda or baru:
        ch.notes.append(f"wip_movements: {beda} beda + {baru} baru vs prod - jalankan ulang "
                        f"bagian 6 dari import_blueprint_2026/import_blueprint.sql")
    else:
        ch.notes.append(f"wip_movements: {ok} baris IMP-2026 di prod sudah = Excel bersih "
                        f"({flag} di antaranya DISCREPANCY_FLAG, setia pada sumber).")
    return rows, {"excel": len(wips), "OK": ok, "BEDA": beda, "BARU": baru,
                  "DISCREPANCY_FLAG": flag, "HANYA_DI_WEB": len(web_only)}, web_only


def rec_cutting(cuts, ch: Changes):
    prod = {c["so_number"]: c for c in load_snap("production.json")["cutting_records"]}
    rows, ok, beda, baru, keputusan = [], 0, 0, 0, 0
    NUMF = ["qty_cut", "main_fabric_used", "puring_used",
            "main_consumption_rate", "puring_consumption_rate"]
    for c in cuts:
        r = clean_row(c, "cutting_records")
        r["cutting_date"] = isod(r["cutting_date"])
        so = r["so_number"]
        p = prod.get(so)
        notes = []
        if not p:
            baru += 1
            notes.append("BARU: cutting_record belum ada di prod")
            ch.cut.append(
                f"INSERT INTO cutting_records (id,so_id,cutting_date,qty_cut,main_fabric_used,"
                f"puring_used,main_consumption_rate,puring_consumption_rate) SELECT "
                f"gen_random_uuid()::text,so.id,{sqlstr(r['cutting_date'])},{int(r['qty_cut'] or 0)},"
                f"{float(r['main_fabric_used'] or 0)},{float(r['puring_used'] or 0)},"
                f"{float(r['main_consumption_rate'] or 0)},{float(r['puring_consumption_rate'] or 0)} "
                f"FROM sales_orders so WHERE so.so_number={sqlstr(so)} AND NOT EXISTS "
                f"(SELECT 1 FROM cutting_records cr WHERE cr.so_id=so.id);")
        else:
            sets, row_keputusan = [], False
            for f in NUMF:
                mode, note = num_delta(f, r[f], p[f], "cutting_records")
                if note:
                    notes.append(note)
                if mode == "excel":
                    sets.append(f"{f}={int(r[f] or 0) if f == 'qty_cut' else float(r[f] or 0)}")
                elif mode == "keepprod":
                    r[f] = float(p[f] or 0)   # tampilkan angka prod di sheet rapi
                    row_keputusan = True
            if sets:
                beda += 1
                ch.cut.append(
                    f"UPDATE cutting_records SET {', '.join(sets)} WHERE so_id=(SELECT id FROM "
                    f"sales_orders WHERE so_number={sqlstr(so)});"
                    f"  -- prod: qty_cut={p['qty_cut']} main_fabric_used={p['main_fabric_used']}")
            elif row_keputusan:
                keputusan += 1
            else:
                ok += 1
        r["_REKONSILIASI"] = " | ".join(notes) if notes else "OK - sama dengan prod"
        rows.append(r)
    excel_so = {c["so_number"] for c in cuts}
    web_only = [("cutting_records", c["so_number"], json.dumps(c, ensure_ascii=False))
                for so, c in sorted(prod.items()) if so not in excel_so]
    seeded = sum(1 for p in prod.values()
                 if approx(p.get("main_consumption_rate"), 1.3)
                 and approx(p.get("puring_consumption_rate"), 0.2))
    if seeded:
        ch.notes.append(
            f"cutting_records: {seeded} baris di prod memakai rate persis 1.3 / 0.2 "
            f"(pola data estimasi/seed - lih. seed_pipeline_data). Untuk baris itu angka "
            f"Excel kemungkinan yang benar; puring_used=0 di Excel tetap ditandai "
            f"PERLU KEPUTUSAN (bisa berarti 'tanpa puring' atau 'belum dicatat').")
    return rows, {"excel": len(cuts), "OK": ok, "BEDA": beda, "BARU": baru,
                  "PERLU_KEPUTUSAN": keputusan, "HANYA_DI_WEB": len(web_only)}, web_only


def rec_reject(rejs, ch: Changes):
    prod = load_snap("production.json")["reject_logs"]
    pindex = defaultdict(list)
    for p in prod:
        pindex[p["so_number"]].append(p)
    rows, ok, beda, baru = [], 0, 0, 0
    matched = set()
    for rj in rejs:
        r = clean_row(rj, "reject_logs")
        so = r["so_number"]
        cand = pindex.get(so, [])
        hit = next((p for i, p in enumerate(cand)
                    if (so, i) not in matched and approx(p["qty_reject"], r["qty_reject"])), None)
        notes = []
        if hit is None:
            hit = next((p for i, p in enumerate(cand) if (so, i) not in matched), None)
        if hit is None:
            baru += 1
            notes.append("BARU: reject_log belum ada di prod")
        else:
            matched.add((so, cand.index(hit)))
            if not approx(hit["qty_reject"], r["qty_reject"]):
                notes.append(f"BEDA qty_reject: excel={r['qty_reject']} prod={hit['qty_reject']}")
                beda += 1
            else:
                ok += 1
        r["_REKONSILIASI"] = " | ".join(notes) if notes else "OK - ada di prod"
        rows.append(r)
    web_only = []
    for so, lst in pindex.items():
        for i, p in enumerate(lst):
            if (so, i) not in matched:
                web_only.append(("reject_logs", so, json.dumps(p, ensure_ascii=False)))
    return rows, {"excel": len(rejs), "OK": ok, "BEDA": beda, "BARU": baru,
                  "HANYA_DI_WEB": len(web_only)}, web_only


def rec_partners(partners, ch: Changes):
    prod = load_snap("fabric.json")["partners"]
    pnames = {norm(p["name"]) for p in prod}
    rows, ok, baru = [], 0, 0
    for pc in partners:
        r = clean_row(pc, "partners")
        if norm(r["name"]) in pnames:
            ok += 1
            note = "OK - sudah ada di prod"
        else:
            baru += 1
            note = "BARU: partner belum ada di prod"
            ch.ins.append(
                f"INSERT INTO partners (id,code,name,category) SELECT gen_random_uuid()::text,"
                f"{sqlstr('IMP-' + r['category'][:3] + '-' + str(abs(hash(r['name'])) % 100000))},"
                f"{sqlstr(r['name'])},{sqlstr(r['category'])} WHERE NOT EXISTS "
                f"(SELECT 1 FROM partners WHERE name={sqlstr(r['name'])});")
        r["_REKONSILIASI"] = note
        rows.append(r)
    excel_names = {norm(p["name"]) for p in partners}
    web_only = [("partners", p["name"], json.dumps(p, ensure_ascii=False))
                for p in prod if norm(p["name"]) not in excel_names]
    return rows, {"excel": len(partners), "OK": ok, "BARU": baru,
                  "HANYA_DI_WEB": len(web_only)}, web_only


# --------------------------------------------------------------------------- #
# Penulis workbook / CSV / laporan
# --------------------------------------------------------------------------- #
HEAD_FILL = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(bold=True, color="FFFFFF")


def _sheet(wb, title, cols, rows):
    ws = wb.create_sheet(title[:31])
    head = cols + ["_REKONSILIASI"]
    ws.append(head)
    for c in range(1, len(head) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
    for r in rows:
        ws.append([r.get(c) for c in head])
    ws.freeze_panes = "A2"
    for i, col in enumerate(head, 1):
        w = max(len(str(col)), *(len(str(r.get(col, ""))) for r in rows)) if rows else len(col)
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 60)
    return ws


def _summary_sheet(wb, summ: dict):
    ws = wb.create_sheet("_RINGKASAN", 0)
    keys = ["excel", "OK", "BEDA", "BARU", "PERLU_KEPUTUSAN", "DISCREPANCY_FLAG", "HANYA_DI_WEB"]
    ws.append(["entitas"] + keys)
    for c in range(1, len(keys) + 2):
        ws.cell(row=1, column=c).fill = HEAD_FILL
        ws.cell(row=1, column=c).font = HEAD_FONT
    for ent, s in summ.items():
        ws.append([ent] + [s.get(k, "") for k in keys])
    ws.column_dimensions["A"].width = 26
    ws.freeze_panes = "B2"


def _weblist_sheet(wb, web_rows):
    ws = wb.create_sheet("HANYA DI WEB")
    ws.append(["entitas", "kunci", "isi baris prod (JSON)"])
    for c in range(1, 4):
        ws.cell(row=1, column=c).fill = HEAD_FILL
        ws.cell(row=1, column=c).font = HEAD_FONT
    for row in web_rows:
        ws.append(list(row))
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 90
    ws.freeze_panes = "A2"


def write_workbook(path: Path, sheets: list, summ: dict, web_rows: list):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for title, ent, rows in sheets:
        _sheet(wb, title, COLS[ent], rows)
    _summary_sheet(wb, summ)
    _weblist_sheet(wb, web_rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def write_csv(path: Path, cols, rows):
    head = cols + ["_REKONSILIASI"]
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=head, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k) for k in head})


def write_report(path: Path, title, summ, web_rows, rows_by_ent, extra):
    L = [f"# {title}", "",
         f"Dibuat: {datetime.now().isoformat(timespec='seconds')}  ",
         "Aturan: **Excel menang** - nilai di workbook rapi = hasil parsing Excel "
         "yang sudah dibersihkan. Selisih vs prod tercatat di `perubahan_prod.sql` "
         "(review manual, tidak dijalankan otomatis). Sel uang kosong di Excel "
         "**tidak** dianggap 0 - kalau prod terisi, angka prod dipertahankan & "
         "ditandai `PERLU KEPUTUSAN`.", "",
         "## Ringkasan per entitas", "",
         "| entitas | baris Excel | OK | BEDA | BARU | perlu keputusan | hanya di web |",
         "|---|--:|--:|--:|--:|--:|--:|"]
    for ent, s in summ.items():
        L.append(f"| {ent} | {s.get('excel','')} | {s.get('OK','')} | {s.get('BEDA','')} "
                 f"| {s.get('BARU','')} | {s.get('PERLU_KEPUTUSAN', s.get('DISCREPANCY_FLAG',''))} "
                 f"| {s.get('HANYA_DI_WEB','')} |")
    L += ["", "## Baris PERLU KEPUTUSAN / BEDA (butuh mata manusia)", ""]
    n = 0
    for ent, rows in rows_by_ent.items():
        flagged = [r for r in rows if not r["_REKONSILIASI"].startswith("OK")]
        if not flagged:
            continue
        L.append(f"### {ent}  ({len(flagged)})")
        for r in flagged[:400]:
            key = r.get("item_code") or r.get("so_number") or r.get("roll_number") \
                or r.get("surat_jalan_no") or r.get("name") or "?"
            L.append(f"- `{key}` - {r['_REKONSILIASI']}")
            n += 1
        if len(flagged) > 400:
            L.append(f"- … ({len(flagged) - 400} lagi, lihat sheet Excel)")
        L.append("")
    if not n:
        L.append("_Tidak ada - semua baris cocok dengan prod._\n")
    L += ["## HANYA DI WEB (ada di prod, tak ada di Excel - TIDAK dihapus)", ""]
    if web_rows:
        cur = None
        for ent, key, _ in web_rows:
            if ent != cur:
                cur = ent
                L.append(f"### {ent}")
            L.append(f"- `{key}`")
        L.append("")
    else:
        L.append("_Tidak ada._\n")
    if extra:
        L += ["## Catatan", ""] + [f"- {e}" for e in extra]
    path.write_text("\n".join(L), encoding="utf-8")


# --------------------------------------------------------------------------- #
def main():
    args = sys.argv[1:]
    f1 = Path(args[0]) if len(args) > 0 else DOWNLOADS / "Salinan dari DATA BAHAN NEW 2026.xlsx"
    f2 = Path(args[1]) if len(args) > 1 else DOWNLOADS / "Salinan dari Monitoring EX PRODUKSI 2026.xlsx"
    for f in (f1, f2):
        if not f.exists():
            sys.exit(f"ERROR: tidak ketemu: {f}")
    OUTDIR.mkdir(parents=True, exist_ok=True)
    ch = Changes()

    # parse KEDUA file dulu - rec_codeso perlu tahu SO mana yang juga ada di Monitoring
    print(f"parse {f1.name}")
    items, receipts, allocs, codeso = parse_bahan(f1, [])
    print(f"parse {f2.name}")
    sos, wips, cuts, rejs, partners = parse_monitoring(f2, [])
    mon_numbers = {o["so_number"] for o in sos}

    # ---------- FASE 1 - DATA BAHAN ----------
    inv_rows, inv_s, inv_w = rec_inventory(items, ch)
    rcp_rows, rcp_s, rcp_w = rec_receipts(receipts, ch)
    alc_rows, alc_s, alc_w = rec_allocs(allocs, ch)
    cso_rows, cso_s, _ = rec_codeso(codeso, ch, mon_numbers)

    bahan_summ = {"inventory_items": inv_s, "material_receipts": rcp_s,
                  "material_allocations": alc_s, "sales_orders (Code So)": cso_s}
    bahan_web = inv_w + rcp_w + alc_w
    bahan_sheets = [
        ("inventory_items", "inventory_items", inv_rows),
        ("material_receipts", "material_receipts", rcp_rows),
        ("material_allocations", "material_allocations", alc_rows),
        ("sales_orders (Code So)", "sales_orders_codeso", cso_rows),
    ]
    wb1 = write_workbook(DOWNLOADS / "DATA BAHAN NEW 2026 - RAPI.xlsx",
                         bahan_sheets, bahan_summ, bahan_web)
    for title, ent, rows in bahan_sheets:
        write_csv(OUTDIR / f"bahan_{ent}.csv", COLS[ent], rows)
    write_report(OUTDIR / "REKONSILIASI_BAHAN.md",
                 "Rekonsiliasi DATA BAHAN NEW 2026 vs Web (prod)",
                 bahan_summ, bahan_web,
                 {k: v for k, v in [("inventory_items", inv_rows),
                                    ("material_receipts", rcp_rows),
                                    ("material_allocations", alc_rows),
                                    ("sales_orders (Code So)", cso_rows)]},
                 [n for n in ch.notes if "receipt" in n or "alloc" in n])

    # ---------- FASE 2 - MONITORING EX PRODUKSI ----------
    so_rows, so_s, so_w = rec_sales_orders(sos, ch, {c["so_number"] for c in codeso})
    wip_rows, wip_s, wip_w = rec_wip(wips, ch)
    cut_rows, cut_s, cut_w = rec_cutting(cuts, ch)
    rej_rows, rej_s, rej_w = rec_reject(rejs, ch)
    par_rows, par_s, par_w = rec_partners(partners, ch)

    prod_summ = {"sales_orders (Monitoring)": so_s, "wip_movements": wip_s,
                 "cutting_records": cut_s, "reject_logs": rej_s, "partners": par_s}
    prod_web = so_w + wip_w + cut_w + rej_w + par_w
    prod_sheets = [
        ("sales_orders", "sales_orders", so_rows),
        ("wip_movements", "wip_movements", wip_rows),
        ("cutting_records", "cutting_records", cut_rows),
        ("reject_logs", "reject_logs", rej_rows),
        ("partners", "partners", par_rows),
    ]
    wb2 = write_workbook(DOWNLOADS / "Monitoring EX PRODUKSI 2026 - RAPI.xlsx",
                         prod_sheets, prod_summ, prod_web)
    for title, ent, rows in prod_sheets:
        write_csv(OUTDIR / f"produksi_{ent}.csv", COLS[ent], rows)
    write_report(OUTDIR / "REKONSILIASI_PRODUKSI.md",
                 "Rekonsiliasi Monitoring EX PRODUKSI 2026 vs Web (prod)",
                 prod_summ, prod_web,
                 {k: v for k, v in [("sales_orders (Monitoring)", so_rows),
                                    ("wip_movements", wip_rows),
                                    ("cutting_records", cut_rows),
                                    ("reject_logs", rej_rows),
                                    ("partners", par_rows)]},
                 [n for n in ch.notes if any(t in n for t in ("wip", "cutting", "reject"))])

    ch.dump(OUTDIR / "perubahan_prod.sql")

    print("\n=== OUTPUT ===")
    for p in (wb1, wb2):
        print(f"  {p}  ({p.stat().st_size:,} B)")
    for p in sorted(OUTDIR.glob("*")):
        if p.is_file():
            print(f"  {p.relative_to(OUTDIR.parent)}  ({p.stat().st_size:,} B)")
    print("\n=== RINGKAS REKONSILIASI ===")
    for grp in (bahan_summ, prod_summ):
        for ent, s in grp.items():
            print(f"  {ent:32s} " + "  ".join(f"{k}={v}" for k, v in s.items()))


if __name__ == "__main__":
    main()
