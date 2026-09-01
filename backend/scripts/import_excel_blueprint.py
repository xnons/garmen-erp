"""
Pembersih & generator impor Excel -> SQL untuk SKEMA BLUEPRINT (6-fase).

Menggantikan scripts/import_excel_bahan.py (yang menyasar skema lama
bahan_baku/log_mutasi_bahan). DB live memakai skema blueprint:
partners / sales_orders / wip_movements / cutting_records / reject_logs /
inventory_items / material_receipts / material_allocations.

Sumber:
  A. "DATA BAHAN NEW 2026.xlsx"
       - "Stok Bahan"    -> inventory_items          (UPSERT on item_code)
       - "Barang Masuk"  -> material_receipts        (marker roll_number 'IMP-%')
       - "Barang Keluar" -> material_allocations     (marker surat_jalan_no 'IMP-%')
       - "Code So"        -> sales_orders            (UPSERT minimal on so_number)
  B. "Monitoring EX PRODUKSI 2026.xlsx"
       - "Monitoring"        -> sales_orders (UPSERT) + wip_movements (marker 'IMP-%')
       - "GUDANG BAHAN"      -> cutting_records      (INSERT bila so belum punya record)
       - "FORM RIJEK BORDIR" -> reject_logs          (stage EMBROIDERY_DEFECT, wip NULL)
       - "DATA"              -> partners             (INSERT WHERE NOT EXISTS by name)

Idempotent: aman dijalankan ulang di Supabase SQL Editor. FK di-resolve saat
run lewat sub-query (so_number / item_code / partner.name) sehingga cocok dengan
baris UUID yang sudah ada.

Pakai:
    python scripts/import_excel_blueprint.py "DATA BAHAN NEW 2026.xlsx" "Monitoring EX PRODUKSI 2026.xlsx" [outdir]
"""
from __future__ import annotations

import csv
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl

MARK = "IMP-2026"                 # penanda baris hasil impor (untuk delete-by-marker)
DEF_DATE = date(2026, 1, 1)

# --------------------------------------------------------------------------- #
# Kanonikalisasi nama rekanan  (selaras dgn Downloads/skrip/sql.py + data DB)
# --------------------------------------------------------------------------- #
BUYER_CANON = {
    "AL-ITIHAD": "AL-ITIHAD", "BINTANG MADANI": "BINTANG MADANI",
    "CAMO": "CAMO WARBROKE", "CAMO WARBROKE": "CAMO WARBROKE",
    "DELUSI": "DELUSI", "FADFAD": "FADFAD", "GOTO FADFAD": "GOTO FADFAD",
    "INSIGHT": "INSIGHT ( SMBU )", "INSIGHT ( SMBU )": "INSIGHT ( SMBU )",
    "PT.SMBU ( INSIGHT )": "INSIGHT ( SMBU )",
    "NEVER SURENDER": "NEVER SURRENDER", "NEVER SURRENDER": "NEVER SURRENDER",
    "NEVERSURENDER": "NEVER SURRENDER", "NVSR": "NEVER SURRENDER",
    "OXFOORD": "OXFOORD", "PA DENNY": "PAKDENNY", "PAKDENNY": "PAKDENNY",
    "PAMOKIDS": "PAMOKIDS",
    "PLANETSUR ( SMBU )": "PLANETSUR ( SMBU )",
    "PT.SMBU ( PLANET SURF )": "PLANETSUR ( SMBU )",
    "PSM": "PSM",
    "PT.SMBU ( JUICE EMATIC )": "JUICE EMATIC ( SMBU )",
    "JUICE EMATIC ( SMBU )": "JUICE EMATIC ( SMBU )",
    "PT.SMBU ( SPYDERBILT )": "SPYDERBILT ( SMBU )",
    "SPYDERBILT ( SMBU )": "SPYDERBILT ( SMBU )",
    "PT.SMBU ( VOXFLY )": "VOXFLY ( SMBU )", "VOXFLY ( SMBU )": "VOXFLY ( SMBU )",
    "PT.SMBU ( PLANET SURF )": "PLANETSUR ( SMBU )",
    "SERAGAM": "SERAGAM", "SERPARANG": "SERPARANG",
    "SEVENTYFOUR": "SEVENTYFOUR", "SGI": "SGI", "SMBU": "SMBU",
    "TRAVEOLOGY": "TRAVEOLOGY", "WARNING": "WARNING",
    "WARNING CLOTHING": "WARNING", "WARNING CLOTIHING": "WARNING",
    "WARNING CLOTIHING ": "WARNING",
    "WILMER": "WILMER STUDIOS", "WILMER STUDIOS": "WILMER STUDIOS",
    "PA ENDOY": "PA ENDOY", "MANAGEMENT": "MANAGEMENT",
}
SUBCON_CANON = {
    "MASTER LAUDRY": "MASTER LAUNDRY", "MASTER LAUNDRY": "MASTER LAUNDRY",
    "CJM EMBORDIERY": "CJM EMBROIDERY", "CJM EMBROIDERY": "CJM EMBROIDERY",
}

# Allowlist rekanan yang sah dari sheet "DATA" (kolomnya bocor ke baris sampah
# di bawah — hanya nama di daftar ini yang di-INSERT).
KNOWN_PARTNERS = {
    "MAKLUN_SEWING": {
        "AL-ITIHAD GARMENT", "PAK ADE SMD", "PAK ADE CIPARAY", "A DADANG",
        "PAK AEP TASIK", "MASTER PA NANA", "PAK ENGKUS", "MASTER PA PIAN", "MANAGEMENT",
    },
    "SUBCON_EMBROIDERY": {"CJM EMBROIDERY", "KO DEDE EMBRO"},
    "SUBCON_WASHING": {
        "RITE CLEAN WASHING", "MASTER LAUNDRY", "ANUGRAH WASHING",
        "BLESSINDO WASHING", "ELPITO WASHING",
    },
    "SUBCON_PRINT": {"PA GANDA PRINT", "MAS KIRNO PRINT", "CIPTA JAYA PRINT"},
}
KNOWN_BUYERS = set(BUYER_CANON.values()) - {"MANAGEMENT"}  # MANAGEMENT bukan buyer

STATUS_MAP = {
    "TERKIRIM": "SHIPPED", "KIRIM": "SHIPPED", "SELESAI": "CLOSED",
    "FINISHING": "FINISHING", "FINISHING ": "FINISHING",
    "WASHING": "WASHING", "BORDIR": "WIP_SUBCON", "EMBRO": "WIP_SUBCON",
    "PRODUKSI": "SEWING", "SEWING": "SEWING", "JAHIT": "SEWING",
    "PERSIAPAN": "CUTTING", "CUTTING": "CUTTING", "POTONG": "CUTTING",
}


# --------------------------------------------------------------------------- #
# Helper pembersihan
# --------------------------------------------------------------------------- #
def s(v) -> str:
    if v is None:
        return ""
    t = str(v).strip()
    if t.upper() in ("#REF!", "#ERROR!", "#N/A", "#VALUE!", "#DIV/0!", "NULL", "-", "."):
        return ""
    return t


def num(v):
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 4)
    t = str(v).strip().replace(",", "")
    if not t or t.startswith("#"):
        return None
    try:
        return round(float(t), 4)
    except ValueError:
        return None


def i2(v):
    """Integer bersih dari sel Excel (angka/teks/None/#REF!) -> int atau None."""
    n = num(v)
    if n is None:
        return None
    return int(round(n))


def to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    t = s(v)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(t[:10], fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def canon_buyer(name: str) -> str:
    n = s(name)
    return BUYER_CANON.get(n.upper(), n)


def canon_partner(name: str) -> str:
    n = s(name)
    return SUBCON_CANON.get(n.upper(), n)


def item_type_of(desc: str) -> str:
    d = desc.upper()
    if "PURING" in d or "JALA" in d:
        return "PURING"
    if "KAIN KERAS" in d or "INTERLINING" in d or "KUMKer".upper() in d or "VISELIN" in d:
        return "INTERLINING"
    if any(k in d for k in ("KANCING", "BENANG", "RESLETING", "ZIPPER", "LABEL", "KARET", "ELASTIC")):
        return "TRIMS_ACCESSORY"
    return "FABRIC_MAIN"


def sqlstr(v) -> str:
    if v is None or v == "":
        return "NULL"
    return "'" + str(v).replace("'", "''") + "'"


def sqlnum(v) -> str:
    return "NULL" if v is None else repr(float(v))


def sqlint(v) -> str:
    return "NULL" if v is None else str(int(v))


def sqldate(d) -> str:
    return "NULL" if not d else f"'{d.isoformat()}'"


def buyer_subq(brand: str) -> str:
    c = canon_buyer(brand)
    if not c:
        return "NULL"
    return f"(SELECT id FROM partners WHERE name = {sqlstr(c)} ORDER BY created_at LIMIT 1)"


def partner_subq(name: str) -> str:
    c = canon_partner(name)
    if not c:
        return "NULL"
    return f"(SELECT id FROM partners WHERE name = {sqlstr(c)} ORDER BY created_at LIMIT 1)"


def so_subq(so: str) -> str:
    return f"(SELECT id FROM sales_orders WHERE so_number = {sqlstr(so)} LIMIT 1)"


def item_subq(code: str) -> str:
    return f"(SELECT id FROM inventory_items WHERE item_code = {sqlstr(code)} LIMIT 1)"


def has_content(r) -> bool:
    return any(c is not None and str(c).strip() != "" for c in r)


# --------------------------------------------------------------------------- #
# PARSER — file A: DATA BAHAN
# --------------------------------------------------------------------------- #
def parse_bahan(path: Path, rep: list):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    # ---- Stok Bahan -> inventory_items (header R5) ----
    ws = wb["Stok Bahan"]
    rows = list(ws.iter_rows(values_only=True))[5:]
    items, seen = [], set()
    neg = noprice = nodesc = 0
    for r in rows:
        code = s(r[3])
        if not code or code in seen:
            continue
        seen.add(code)
        desc = s(r[4]) or f"(TANPA DESKRIPSI) {code}"
        if not s(r[4]):
            nodesc += 1
        stok = num(r[14]) or 0.0
        if stok < 0:
            neg += 1
            stok = 0.0
        price = num(r[16]) or 0.0
        if price == 0.0:
            noprice += 1
        items.append({
            "item_code": code,
            "description": desc,
            "item_type": item_type_of(desc),
            "unit": s(r[9]) or "YARD",
            "unit_price": round(price, 2),
            "current_stock": round(stok, 2),
            "min_stock_alert": 50.0,
            "width_inch": 58.0,
            "rack_location": "GUDANG_UTAMA",
            "_brand": canon_buyer(r[1]) or canon_buyer(r[5]) or "-",
            "_status": s(r[7]).upper() or "FOB",
            "_tanggal": (to_date(r[2]) or to_date(r[6])),
        })
    rep.append(f"### inventory_items — {len(items)} item")
    rep.append(f"- deskripsi kosong -> placeholder: {nodesc}")
    rep.append(f"- stok negatif (noise) -> 0: {neg}")
    rep.append(f"- item harga 0 (bahan CMT milik customer): {noprice}")

    valid_codes = {i["item_code"] for i in items}

    # ---- Barang Masuk -> material_receipts (header R4) ----
    ws = wb["Barang Masuk"]
    rows = list(ws.iter_rows(values_only=True))[4:]
    receipts = []
    seq = {}
    skc = skq = 0
    for r in rows:
        if not has_content(r):
            continue
        code = s(r[0])
        if not code:
            skc += 1
            continue
        if code not in valid_codes:
            skc += 1
            continue
        qty = num(r[2])
        if qty is None or qty <= 0:
            skq += 1
            continue
        seq[code] = seq.get(code, 0) + 1
        receipts.append({
            "item_code": code,
            "receipt_date": to_date(r[6]) or DEF_DATE,
            "roll_number": f"{MARK}-{code}-{seq[code]:03d}",
            "qty_received": round(qty, 2),
            "unit": s(r[3]) or "YARD",
            "contract_type": (s(r[7]).upper() or "FOB")[:20],
            "inspection_status": "PASSED",
            "_color": s(r[4]),
        })
    rep.append(f"### material_receipts — {len(receipts)} baris  (dilewati: kode {skc}, qty<=0 {skq})")

    # ---- Barang Keluar -> material_allocations (header R4) ----
    ws = wb["Barang Keluar"]
    rows = list(ws.iter_rows(values_only=True))[4:]
    allocs = []
    skc = skq = skso = 0
    for r in rows:
        if not has_content(r):
            continue
        code = s(r[1])
        if not code or code not in valid_codes:
            skc += 1
            continue
        qty = num(r[3])
        if qty is None or qty <= 0:
            skq += 1
            continue
        so = s(r[6])
        if not so or not so.upper().startswith("SO"):
            skso += 1
            continue
        allocs.append({
            "so_number": so,
            "item_code": code,
            "dispatch_date": to_date(r[0]) or DEF_DATE,
            "qty_issued": round(qty, 2),
            "surat_jalan_no": f"{MARK}-{so}-{code}",
            "_warna": s(r[5]),
        })
    rep.append(f"### material_allocations — {len(allocs)} baris  "
               f"(dilewati: kode {skc}, qty<=0 {skq}, tanpa SO {skso})")

    # ---- Code So -> sales_orders minimal (header R3) ----
    ws = wb["Code So"]
    rows = list(ws.iter_rows(values_only=True))[3:]
    codeso, seen = [], set()
    for r in rows:
        so = s(r[0])
        if not so or not so.upper().startswith("SO") or so in seen:
            continue
        seen.add(so)
        codeso.append({
            "so_number": so,
            "brand": canon_buyer(r[1]),
            "style_name": s(r[2]) or so,
        })
    rep.append(f"### Code So — {len(codeso)} SO (prefill sales_orders)")
    wb.close()
    return items, receipts, allocs, codeso


# --------------------------------------------------------------------------- #
# PARSER — file B: MONITORING EX PRODUKSI
# --------------------------------------------------------------------------- #
WIP_STAGES = [
    # (stage_name, sequence_order, kolom qty, vendor?)
    ("PRINT_MENTAH", 1, 17, False),
    ("EMBROIDERY_MENTAH", 2, 16, False),
    ("SEWING_MAKLUN", 3, 18, True),   # dispatched=18 Prod, received=19 Hasil prod
    ("WASHING", 4, 22, False),
    ("EMBROIDERY_JADI", 5, 21, False),
    ("FINISHING", 6, 23, False),
]


def parse_monitoring(path: Path, rep: list):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    # ---- Monitoring -> sales_orders + wip_movements (header R5) ----
    ws = wb["Monitoring"]
    rows = list(ws.iter_rows(values_only=True))[5:]
    sos, wips = [], []
    seen = set()
    for r in rows:
        if not has_content(r) or len(r) < 12:
            continue
        so = s(r[2])
        if not so or not so.upper().startswith("SO") or so in seen:
            continue
        seen.add(so)
        odate = to_date(r[1]) or DEF_DATE
        udate = to_date(r[9]) or odate
        order_qty = i2(r[8]) or 0
        status = STATUS_MAP.get(s(r[10]).upper(), "REGISTERED")
        sos.append({
            "so_number": so,
            "brand": canon_buyer(r[3]),
            "style_name": s(r[4]) or so,
            "item_category": (s(r[5]) or "GARMENT")[:100],
            "color": s(r[7])[:50] or None,
            "order_qty": order_qty,
            "status": status,
            "order_date": odate,
            "special_instructions": (f"Model: {s(r[6])}" if s(r[6]) else None),
        })
        vendor = s(r[11])
        rej_total = i2(r[26]) or 0
        if rej_total < 0:
            rej_total = 0
        for stage, seqn, qcol, is_vendor in WIP_STAGES:
            q = i2(r[qcol]) if qcol < len(r) else None
            if not q or q <= 0:
                continue
            if stage == "SEWING_MAKLUN":
                disp = q
                recv = i2(r[19]) if 19 < len(r) else q
                recv = recv if recv is not None else q
                rej = min(rej_total, disp) if rej_total else 0
                bal = disp - (recv + rej)
                st = ("DISCREPANCY_FLAG" if bal != 0
                      else "COMPLETED" if recv >= disp else "PARTIAL_RECEIVED")
            else:
                disp = recv = q
                rej = 0
                bal = 0
                st = "COMPLETED"
            wips.append({
                "so_number": so,
                "stage_name": stage,
                "sequence_order": seqn,
                "vendor": vendor if is_vendor else "",
                "surat_jalan_no": f"{MARK}-{so}-{stage}",
                "dispatch_date": udate,
                "qty_dispatched": disp,
                "received_date": udate,
                "qty_received": recv,
                "qty_reject": rej,
                "balance_discrepancy": bal,
                "status": st,
            })
    rep.append(f"### sales_orders (Monitoring) — {len(sos)} SO")
    rep.append(f"### wip_movements — {len(wips)} baris  "
               f"(snapshot: tanggal dispatch=terima diambil dari 'Tgl Update'; "
               f"qty dispatch≈terima utk tahap non-sewing; penanda surat_jalan_no '{MARK}-%')")

    # ---- GUDANG BAHAN -> cutting_records (header R5) ----
    ws = wb["GUDANG BAHAN"]
    rows = list(ws.iter_rows(values_only=True))[5:]
    cuts, seen = [], set()
    skso = 0
    for r in rows:
        if not has_content(r):
            continue
        so = s(r[2]) or s(r[13]) or s(r[26])
        if not so or not so.upper().startswith("SO"):
            skso += 1
            continue
        if so in seen:
            continue
        seen.add(so)
        qty_cut = i2(r[8]) or i2(r[34]) or 0
        if qty_cut <= 0:
            continue
        rate_main = num(r[35]) or num(r[15]) or 0.0
        fabric_main = num(r[33]) or (round(qty_cut * rate_main, 2) if rate_main else 0.0)
        puring_qty = num(r[39]) if len(r) > 39 else None
        rate_pur = num(r[17]) or 0.0
        cuts.append({
            "so_number": so,
            "cutting_date": to_date(r[1]) or DEF_DATE,
            "qty_cut": qty_cut,
            "main_fabric_used": round(fabric_main or 0.0, 2),
            "puring_used": round(puring_qty or 0.0, 2),
            "main_consumption_rate": round(rate_main or 0.0, 4),
            "puring_consumption_rate": round(rate_pur or 0.0, 4),
            "_jenis_bahan": s(r[30]) if len(r) > 30 else "",
            "_warna": s(r[31]) if len(r) > 31 else "",
        })
    rep.append(f"### cutting_records — {len(cuts)} SO  (dilewati tanpa SO: {skso}; "
               f"1 record per SO, hanya bila SO belum punya cutting_record)")

    # ---- FORM RIJEK BORDIR -> reject_logs (header R9) ----
    ws = wb["FORM RIJEK BORDIR"]
    rows = list(ws.iter_rows(values_only=True))[9:]
    rejs = []
    skso = skq = 0
    for r in rows:
        if not has_content(r):
            continue
        so = s(r[2])
        if not so or not so.upper().startswith("SO"):
            skso += 1
            continue
        qty = i2(r[8])
        if not qty or qty <= 0:
            skq += 1
            continue
        harga = num(r[7]) or 0.0
        total = num(r[9]) or round(qty * harga, 2)
        rejs.append({
            "so_number": so,
            "stage_name": "EMBROIDERY_DEFECT",
            "defect_reason": ("CACAT BORDIR" + (f" - {s(r[6])}" if s(r[6]) else ""))[:100],
            "qty_reject": qty,
            "unit_cost_loss": round(harga, 2),
            "total_loss": round(total, 2),
        })
    rep.append(f"### reject_logs — {len(rejs)} baris  "
               f"(dilewati: tanpa SO {skso}, qty<=0 {skq}; stage EMBROIDERY_DEFECT)")

    # ---- DATA -> partners (header R4) ----
    ws = wb["DATA"]
    rows = list(ws.iter_rows(values_only=True))[4:]
    part_cols = [(2, "BUYER"), (6, "MAKLUN_SEWING"), (8, "SUBCON_EMBROIDERY"),
                 (10, "SUBCON_WASHING"), (12, "SUBCON_PRINT")]
    partners, pseen = [], set()
    for r in rows:
        for idx, cat in part_cols:
            if idx >= len(r):
                continue
            raw = s(r[idx])
            if not raw:
                continue
            name = canon_buyer(raw) if cat == "BUYER" else canon_partner(raw)
            if not name:
                continue
            ok = name in KNOWN_BUYERS if cat == "BUYER" else name in KNOWN_PARTNERS.get(cat, set())
            if not ok:
                continue  # baris sampah dari kolom "DATA" yang bocor
            key = (name.upper(), cat)
            if key in pseen:
                continue
            pseen.add(key)
            partners.append({"name": name, "category": cat})
    rep.append(f"### partners (DATA) — {len(partners)} kandidat (hanya di-INSERT bila nama belum ada)")
    wb.close()
    return sos, wips, cuts, rejs, partners


# --------------------------------------------------------------------------- #
# WRITER
# --------------------------------------------------------------------------- #
def write_csv(path: Path, rows: list):
    if not rows:
        return
    keys = list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=keys)
        w.writeheader()
        w.writerows(rows)


def write_sql(path: Path, data: dict):
    items = data["items"]; receipts = data["receipts"]; allocs = data["allocs"]
    codeso = data["codeso"]; sos = data["sos"]; wips = data["wips"]
    cuts = data["cuts"]; rejs = data["rejs"]; partners = data["partners"]

    L = []
    ap = L.append
    ap("-- =====================================================================")
    ap("-- IMPOR EXCEL -> SKEMA BLUEPRINT (generated by import_excel_blueprint.py)")
    ap(f"-- Dibuat: {datetime.now().isoformat(timespec='seconds')}")
    ap("-- Target: PostgreSQL / Supabase SQL Editor. Idempotent (UPSERT + marker).")
    ap("-- =====================================================================")
    ap("BEGIN;")
    ap("")

    # 1. partners (INSERT WHERE NOT EXISTS by name)
    ap(f"-- 1. partners  ({len(partners)} kandidat) --------------------------------")
    for p in partners:
        code = f"IMP-{p['category'][:3]}-{abs(hash(p['name'])) % 100000:05d}"
        ap(f"INSERT INTO partners (id, code, name, category) "
           f"SELECT gen_random_uuid()::text, {sqlstr(code)}, {sqlstr(p['name'])}, {sqlstr(p['category'])} "
           f"WHERE NOT EXISTS (SELECT 1 FROM partners WHERE name = {sqlstr(p['name'])});")
    ap("")

    # 2. inventory_items (UPSERT on item_code)
    ap(f"-- 2. inventory_items  ({len(items)}) ------------------------------------")
    for it in items:
        ap("INSERT INTO inventory_items (id, item_code, description, item_type, unit, "
           "unit_price, current_stock, min_stock_alert, width_inch, gramasi_gsm, rack_location) VALUES ("
           f"gen_random_uuid()::text, {sqlstr(it['item_code'])}, {sqlstr(it['description'])}, "
           f"{sqlstr(it['item_type'])}, {sqlstr(it['unit'])}, {sqlnum(it['unit_price'])}, "
           f"{sqlnum(it['current_stock'])}, {sqlnum(it['min_stock_alert'])}, {sqlnum(it['width_inch'])}, "
           f"0, {sqlstr(it['rack_location'])})")
        ap("  ON CONFLICT (item_code) DO UPDATE SET description=EXCLUDED.description, "
           "item_type=EXCLUDED.item_type, unit=EXCLUDED.unit, "
           "unit_price=CASE WHEN EXCLUDED.unit_price>0 THEN EXCLUDED.unit_price ELSE inventory_items.unit_price END, "
           "current_stock=EXCLUDED.current_stock;")
    ap("")

    # 3. sales_orders (UPSERT on so_number) — Code So dulu (minimal), lalu Monitoring (kaya)
    ap(f"-- 3. sales_orders  (Code So {len(codeso)} + Monitoring {len(sos)}) --------")
    for c in codeso:
        ap("INSERT INTO sales_orders (id, so_number, buyer_id, style_name, item_category, order_qty, status, order_date, contract_type) VALUES ("
           f"gen_random_uuid()::text, {sqlstr(c['so_number'])}, {buyer_subq(c['brand'])}, "
           f"{sqlstr(c['style_name'][:150])}, 'GARMENT', 0, 'REGISTERED', {sqldate(DEF_DATE)}, 'CMT')")
        ap("  ON CONFLICT (so_number) DO UPDATE SET "
           "style_name=EXCLUDED.style_name, "
           "buyer_id=COALESCE(sales_orders.buyer_id, EXCLUDED.buyer_id);")
    for o in sos:
        ap("INSERT INTO sales_orders (id, so_number, buyer_id, style_name, item_category, color, "
           "order_qty, status, order_date, special_instructions, contract_type) VALUES ("
           f"gen_random_uuid()::text, {sqlstr(o['so_number'])}, {buyer_subq(o['brand'])}, "
           f"{sqlstr(o['style_name'][:150])}, {sqlstr(o['item_category'])}, {sqlstr(o['color'])}, "
           f"{sqlint(o['order_qty'])}, {sqlstr(o['status'])}, {sqldate(o['order_date'])}, "
           f"{sqlstr(o['special_instructions'])}, 'CMT')")
        ap("  ON CONFLICT (so_number) DO UPDATE SET "
           "style_name=EXCLUDED.style_name, item_category=EXCLUDED.item_category, "
           "color=COALESCE(EXCLUDED.color, sales_orders.color), "
           "order_qty=CASE WHEN EXCLUDED.order_qty>0 THEN EXCLUDED.order_qty ELSE sales_orders.order_qty END, "
           "status=EXCLUDED.status, order_date=EXCLUDED.order_date, "
           "buyer_id=COALESCE(sales_orders.buyer_id, EXCLUDED.buyer_id), "
           "special_instructions=COALESCE(EXCLUDED.special_instructions, sales_orders.special_instructions);")
    ap("")

    # 4. material_receipts (marker)
    ap(f"-- 4. material_receipts  ({len(receipts)}) -------------------------------")
    ap(f"DELETE FROM material_receipts WHERE roll_number LIKE '{MARK}-%';")
    for r in receipts:
        ap("INSERT INTO material_receipts (id, item_id, receipt_date, roll_number, qty_received, unit, contract_type, inspection_status) VALUES ("
           f"gen_random_uuid()::text, {item_subq(r['item_code'])}, {sqldate(r['receipt_date'])}, "
           f"{sqlstr(r['roll_number'])}, {sqlnum(r['qty_received'])}, {sqlstr(r['unit'])}, "
           f"{sqlstr(r['contract_type'])}, {sqlstr(r['inspection_status'])});")
    ap("")

    # 5. material_allocations (marker)
    ap(f"-- 5. material_allocations  ({len(allocs)}) ------------------------------")
    ap(f"DELETE FROM material_allocations WHERE surat_jalan_no LIKE '{MARK}-%';")
    skipped_alloc = 0
    for a in allocs:
        # butuh so_id & item_id ada; kalau SO tak ketemu, INSERT ... SELECT akan 0 baris
        ap("INSERT INTO material_allocations (id, so_id, item_id, dispatch_date, qty_issued, surat_jalan_no) "
           f"SELECT gen_random_uuid()::text, so.id, it.id, {sqldate(a['dispatch_date'])}, "
           f"{sqlnum(a['qty_issued'])}, {sqlstr(a['surat_jalan_no'])} "
           f"FROM sales_orders so, inventory_items it "
           f"WHERE so.so_number = {sqlstr(a['so_number'])} AND it.item_code = {sqlstr(a['item_code'])};")
    ap("")

    # 6. wip_movements (marker)
    ap(f"-- 6. wip_movements  ({len(wips)}) ---------------------------------------")
    ap(f"DELETE FROM wip_movements WHERE surat_jalan_no LIKE '{MARK}-%';")
    for w in wips:
        pid = partner_subq(w["vendor"]) if w["vendor"] else "NULL"
        ap("INSERT INTO wip_movements (id, so_id, stage_name, sequence_order, partner_id, surat_jalan_no, "
           "dispatch_date, qty_dispatched, received_date, qty_received, qty_reject, balance_discrepancy, status) "
           f"SELECT gen_random_uuid()::text, so.id, {sqlstr(w['stage_name'])}, {w['sequence_order']}, "
           f"{pid}, {sqlstr(w['surat_jalan_no'])}, {sqldate(w['dispatch_date'])}, {sqlint(w['qty_dispatched'])}, "
           f"{sqldate(w['received_date'])}, {sqlint(w['qty_received'])}, {sqlint(w['qty_reject'])}, "
           f"{sqlint(w['balance_discrepancy'])}, {sqlstr(w['status'])} "
           f"FROM sales_orders so WHERE so.so_number = {sqlstr(w['so_number'])};")
    ap("")

    # 7. cutting_records (INSERT bila SO belum punya record)
    ap(f"-- 7. cutting_records  ({len(cuts)}) ------------------------------------")
    for c in cuts:
        ap("INSERT INTO cutting_records (id, so_id, cutting_date, qty_cut, main_fabric_used, puring_used, "
           "main_consumption_rate, puring_consumption_rate) "
           f"SELECT gen_random_uuid()::text, so.id, {sqldate(c['cutting_date'])}, {sqlint(c['qty_cut'])}, "
           f"{sqlnum(c['main_fabric_used'])}, {sqlnum(c['puring_used'])}, "
           f"{sqlnum(c['main_consumption_rate'])}, {sqlnum(c['puring_consumption_rate'])} "
           f"FROM sales_orders so WHERE so.so_number = {sqlstr(c['so_number'])} "
           f"AND NOT EXISTS (SELECT 1 FROM cutting_records cr WHERE cr.so_id = so.id);")
    ap("")

    # 8. reject_logs (hapus impor sebelumnya: stage EMBROIDERY_DEFECT tanpa wip)
    ap(f"-- 8. reject_logs  ({len(rejs)}) ---------------------------------------")
    ap("DELETE FROM reject_logs WHERE stage_name = 'EMBROIDERY_DEFECT' AND wip_movement_id IS NULL;")
    for r in rejs:
        ap("INSERT INTO reject_logs (id, so_id, stage_name, defect_reason, qty_reject, unit_cost_loss, total_loss) "
           f"SELECT gen_random_uuid()::text, so.id, {sqlstr(r['stage_name'])}, {sqlstr(r['defect_reason'])}, "
           f"{sqlint(r['qty_reject'])}, {sqlnum(r['unit_cost_loss'])}, {sqlnum(r['total_loss'])} "
           f"FROM sales_orders so WHERE so.so_number = {sqlstr(r['so_number'])};")
    ap("")
    ap("COMMIT;")
    ap("")
    path.write_text("\n".join(L), encoding="utf-8")


# --------------------------------------------------------------------------- #
def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    fa, fb = Path(sys.argv[1]), Path(sys.argv[2])
    outdir = Path(sys.argv[3]) if len(sys.argv) > 3 else \
        Path(__file__).resolve().parent.parent / "database" / "import_blueprint_2026"
    outdir.mkdir(parents=True, exist_ok=True)

    rep = [f"# Laporan Impor Excel -> Skema Blueprint",
           f"", f"Sumber A: `{fa.name}`  ", f"Sumber B: `{fb.name}`  ",
           f"Dibuat: {datetime.now().isoformat(timespec='seconds')}", ""]

    print(f"Baca A: {fa}")
    items, receipts, allocs, codeso = parse_bahan(fa, rep)
    print(f"Baca B: {fb}")
    sos, wips, cuts, rejs, partners = parse_monitoring(fb, rep)

    data = dict(items=items, receipts=receipts, allocs=allocs, codeso=codeso,
                sos=sos, wips=wips, cuts=cuts, rejs=rejs, partners=partners)

    rep += [
        "",
        "## Catatan penting sebelum menjalankan",
        "",
        "1. **Prasyarat**: tabel skema blueprint sudah ada (backend blueprint sudah "
        "jalan/di-deploy). SQL ini tidak membuat tabel, hanya isi data.",
        "2. **Idempotent**: `partners`/`inventory_items`/`sales_orders` pakai UPSERT; "
        "`material_receipts`/`material_allocations`/`wip_movements` dihapus dulu "
        f"berdasarkan penanda `{MARK}-%` lalu diisi ulang; `reject_logs` hapus stage "
        "`EMBROIDERY_DEFECT` tanpa wip; `cutting_records` hanya di-INSERT bila SO "
        "belum punya record. Data manual kamu (penanda lain) tidak tersentuh.",
        "3. **FK buyer**: `sales_orders.buyer_id` di-isi via sub-query nama partner. "
        "Jalankan blok `partners` dulu (sudah diurutkan) supaya semua brand ketemu.",
        "4. **`wip_movements` = snapshot, bukan log gerakan asli.** Sheet Monitoring "
        "hanya menyimpan angka status per tahap, jadi: tanggal dispatch = tanggal "
        "terima = 'Tgl Update'; untuk tahap non-sewing qty dispatch = qty terima; "
        "reject tahap sewing diambil dari kolom 'Total Rijek'. Angka mentah "
        "diterjemahkan apa adanya (termasuk selisih janggal di sumber).",
        "5. **`cutting_records`**: 1 baris per SO dari sheet GUDANG BAHAN. Bila kolom "
        "konsumsi kosong di sumber, `main_fabric_used`/rate = 0.",
        "6. Sheet yang TIDAK diimpor: REKAPAN, Gudang Dewi (semua `#REF!`), "
        "Pengajaun Akse, FORM WI, finishing, dan sheet surat-jalan lainnya.",
    ]

    write_csv(outdir / "01_partners.csv", partners)
    write_csv(outdir / "02_inventory_items.csv", items)
    write_csv(outdir / "03_material_receipts.csv", receipts)
    write_csv(outdir / "04_material_allocations.csv", allocs)
    write_csv(outdir / "05_sales_orders.csv", sos)
    write_csv(outdir / "06_wip_movements.csv", wips)
    write_csv(outdir / "07_cutting_records.csv", cuts)
    write_csv(outdir / "08_reject_logs.csv", rejs)
    write_sql(outdir / "import_blueprint.sql", data)
    (outdir / "LAPORAN_PEMBERSIHAN.md").write_text("\n".join(rep), encoding="utf-8")

    print(f"\nTulis: {outdir}")
    for p in sorted(outdir.iterdir()):
        print(f"  - {p.name}  ({p.stat().st_size:,} B)")
    print(f"\nRingkas: {len(partners)} partners | {len(items)} items | {len(receipts)} receipts | "
          f"{len(allocs)} allocs | {len(sos)} SO | {len(wips)} wip | {len(cuts)} cutting | {len(rejs)} reject")


if __name__ == "__main__":
    main()
