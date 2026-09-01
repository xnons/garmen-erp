"""
Pembersih & generator impor data BAHAN dari Excel -> CSV bersih + SQL.

Sumber : "DATA BAHAN NEW 2026.xlsx" (ekspor Google Sheets, agak berantakan)
Sheet yang dipakai:
  - "Stok Bahan"     -> tabel bahan_baku        (master, 126 item)
  - "Barang Masuk"   -> tabel log_mutasi_bahan  (tipe MASUK)
  - "Barang Keluar"  -> tabel log_mutasi_bahan  (tipe KELUAR_PRODUKSI)
  - "Code So"         -> tabel katalog_so        (master ringan, tabel baru)

Sheet yang DIABAIKAN: REKAPAN (pivot rumus), FABRIC INSPECTION (formulir QC),
Sheet25 (template surat jalan kosong).

Cara pakai:
    python scripts/import_excel_bahan.py "C:/path/DATA BAHAN NEW 2026.xlsx" [outdir]

Output (default: backend/database/import_bahan_2026/):
    01_bahan_baku.csv
    02_log_mutasi_bahan.csv
    03_katalog_so.csv
    import_all.sql
    LAPORAN_PEMBERSIHAN.md

SQL bersifat idempotent: bahan_baku & katalog_so pakai UPSERT (ON CONFLICT),
log_mutasi_bahan dihapus dulu berdasarkan petugas='Import Excel 2026' lalu
di-insert ulang. Aman dijalankan berkali-kali di Supabase SQL Editor.
"""
from __future__ import annotations

import csv
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl

PETUGAS_TAG = "Import Excel 2026"
LOKASI_GUDANG = "Gudang Master Garment (CJM)"

# --------------------------------------------------------------------------- #
# Helper pembersihan
# --------------------------------------------------------------------------- #
def s(v) -> str:
    """String bersih: strip, buang #ERROR!/#N/A, None -> ''."""
    if v is None:
        return ""
    t = str(v).strip()
    if t.upper() in ("#ERROR!", "#N/A", "#VALUE!", "#REF!", "#DIV/0!", "NULL", "-"):
        return ""
    return t


def num(v):
    """Ambil angka; teks rumus error / kosong -> None. Bulatkan noise float."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    t = str(v).strip().replace(",", "")
    if not t or t.startswith("#"):
        return None
    try:
        return round(float(t), 2)
    except ValueError:
        return None


def to_date(v):
    """datetime/tanggal -> date; string 'YYYY-MM-DD' -> date; lainnya -> None."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    t = s(v)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(t[:10], fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def sqlstr(v) -> str:
    """Literal SQL: NULL bila kosong, else '...' dengan escape kutip."""
    if v is None or v == "":
        return "NULL"
    return "'" + str(v).replace("'", "''") + "'"


def sqlnum(v) -> str:
    return "NULL" if v is None else repr(float(v))


# --------------------------------------------------------------------------- #
# 1. Stok Bahan -> bahan_baku
# --------------------------------------------------------------------------- #
def parse_bahan(wb, report):
    ws = wb["Stok Bahan"]
    rows = list(ws.iter_rows(values_only=True))
    # header di baris 5 (index 4); data mulai baris 6
    data = rows[5:]
    out = []
    seen = set()
    skipped_blank = 0
    fixed_status, fixed_negstok, fixed_noprice, fixed_nodesc, fixed_nobrand = 0, 0, 0, 0, 0

    for r in data:
        code = s(r[3])
        if not code:
            skipped_blank += 1
            continue
        if code in seen:
            report.append(f"- ⚠️  Kode ganda di Stok Bahan dilewati: `{code}`")
            continue
        seen.add(code)

        desc = s(r[4])
        if not desc:
            desc = f"(BELUM ADA DESKRIPSI) {code}"
            fixed_nodesc += 1

        status_raw = s(r[7]).upper()
        kategori = status_raw if status_raw in ("FOB", "CMT", "CASH") else ""
        if status_raw and not kategori:
            report.append(f"- ⚠️  STATUS tak dikenal `{r[7]!r}` di {code} -> 'LAINNYA'")
        if not kategori:
            kategori = "LAINNYA"
            if status_raw != "":
                pass
        if s(r[7]) != status_raw and s(r[7]):
            fixed_status += 1

        supplier = s(r[1]) or s(r[5])
        if not supplier:
            supplier = "-"
            fixed_nobrand += 1

        satuan = s(r[9]) or "YARD"

        stok = num(r[14])
        if stok is None:
            stok = 0.0
        if stok < 0:
            report.append(f"- 🔧 Stok negatif {stok} di {code} -> 0 (noise pembulatan)")
            stok = 0.0
            fixed_negstok += 1

        harga = num(r[16]) or 0.0
        if harga == 0.0:
            fixed_noprice += 1

        tgl = to_date(r[2]) or to_date(r[6])
        keterangan = s(r[20]) if len(r) > 20 else ""

        out.append({
            "id": code,
            "kode_sku": code,
            "nama_item": desc,
            "kategori": kategori,
            "satuan": satuan,
            "stok_saat_ini": stok,
            "stok_minimum": 10.0,
            "harga_per_satuan": harga,
            "lokasi_gudang": LOKASI_GUDANG,
            "supplier_utama": supplier,
            "warna_kode": "#3b82f6",
            "no_faktur_po": "-",
            "tanggal_masuk": tgl.isoformat() if tgl else "",
            "tipe_pembayaran": "CASH",
            "status_pembayaran": "LUNAS",
            "jatuh_tempo": "",
            "_keterangan_excel": keterangan,
        })

    report.append("")
    report.append(f"### bahan_baku — {len(out)} item")
    report.append(f"- Baris kosong/spacer dilewati: {skipped_blank}")
    report.append(f"- Deskripsi kosong diisi placeholder: {fixed_nodesc}")
    report.append(f"- STATUS di-trim/normalisasi: {fixed_status}")
    report.append(f"- Brand kosong -> '-': {fixed_nobrand}")
    report.append(f"- Stok negatif dinolkan: {fixed_negstok}")
    report.append(f"- Item tanpa harga (harga=0): {fixed_noprice} (wajar utk bahan CMT milik customer)")
    return out


# --------------------------------------------------------------------------- #
# 2. Barang Masuk + Barang Keluar -> log_mutasi_bahan (dengan saldo berjalan)
# --------------------------------------------------------------------------- #
def parse_mutasi(wb, valid_codes, report):
    events = []  # (bahan_id, tanggal|None, urut_prioritas, seq, tipe, jumlah, ref, catatan)

    def has_content(r):
        return any(c is not None and str(c).strip() != "" for c in r)

    # --- Barang Masuk (header baris 4 / index 3) ---
    ws = wb["Barang Masuk"]
    rows = list(ws.iter_rows(values_only=True))[4:]
    bm_skip_code, bm_skip_qty, bm_skip_fk = 0, 0, 0
    for i, r in enumerate(rows):
        if not has_content(r):
            continue
        code = s(r[0])
        if not code:
            bm_skip_code += 1
            continue
        if code not in valid_codes:
            bm_skip_fk += 1
            report.append(f"- ⚠️  MASUK: kode `{code}` tidak ada di master -> dilewati")
            continue
        qty = num(r[2])
        if qty is None or qty <= 0:
            bm_skip_qty += 1
            continue
        warna = s(r[4])
        ket = s(r[8]) if len(r) > 8 else ""
        catatan = " | ".join(x for x in [f"Warna: {warna}" if warna else "", ket] if x)[:250]
        events.append([code, to_date(r[6]), 0, i, "MASUK", qty, "-", catatan])

    # --- Barang Keluar (header baris 4 / index 3) ---
    ws = wb["Barang Keluar"]
    rows = list(ws.iter_rows(values_only=True))[4:]
    bk_skip_code, bk_skip_qty, bk_skip_fk = 0, 0, 0
    for i, r in enumerate(rows):
        if not has_content(r):
            continue
        code = s(r[1])
        if not code:
            bk_skip_code += 1
            continue
        if code not in valid_codes:
            bk_skip_fk += 1
            report.append(f"- ⚠️  KELUAR: kode `{code}` tidak ada di master -> dilewati")
            continue
        qty = num(r[3])
        if qty is None or qty <= 0:
            bk_skip_qty += 1
            continue
        warna = s(r[5])
        so = s(r[6])
        style = s(r[8]) if len(r) > 8 else ""
        ket = s(r[11]) if len(r) > 11 else ""
        catatan = " | ".join(x for x in [
            f"Warna: {warna}" if warna else "",
            f"Style: {style}" if style else "",
            ket,
        ] if x)[:250]
        events.append([code, to_date(r[0]), 1, i, "KELUAR_PRODUKSI", qty, so or "-", catatan])

    # --- Hitung saldo berjalan per bahan_id ---
    events.sort(key=lambda e: (e[0], e[1] or date.min, e[2], e[3]))
    saldo = {}
    out = []
    for code, tgl, _pri, _seq, tipe, qty, ref, catatan in events:
        before = round(saldo.get(code, 0.0), 2)
        after = round(before + qty, 2) if tipe == "MASUK" else round(before - qty, 2)
        saldo[code] = after
        out.append({
            "bahan_id": code,
            "tanggal": (tgl or date(2026, 1, 1)).isoformat() + " 00:00:00",
            "tipe": tipe,
            "jumlah": qty,
            "stok_sebelum": before,
            "stok_sesudah": after,
            "referensi_po_spk": ref,
            "catatan": catatan,
            "petugas": PETUGAS_TAG,
        })

    report.append("")
    report.append(f"### log_mutasi_bahan — {len(out)} baris")
    report.append(f"- MASUK: dilewati (kode kosong {bm_skip_code}, qty<=0 {bm_skip_qty}, kode asing {bm_skip_fk})")
    report.append(f"- KELUAR: dilewati (kode kosong {bk_skip_code}, qty<=0 {bk_skip_qty}, kode asing {bk_skip_fk})")
    report.append("- Kolom stok_sebelum/stok_sesudah dihitung ulang sebagai saldo berjalan "
                  "per kode (urut tanggal, MASUK sebelum KELUAR di tanggal sama).")
    neg = sorted({r["bahan_id"] for r in out if r["stok_sesudah"] < -0.5})
    if neg:
        report.append(f"- ℹ️  Saldo berjalan sempat minus di {len(neg)} kode "
                      f"(keluar tercatat > masuk di Excel): {', '.join(neg[:15])}"
                      + (" ..." if len(neg) > 15 else ""))
    return out


# --------------------------------------------------------------------------- #
# 3. Code So -> katalog_so
# --------------------------------------------------------------------------- #
def parse_katalog(wb, report):
    ws = wb["Code So"]
    rows = list(ws.iter_rows(values_only=True))[3:]  # header baris 3
    out, seen = [], set()
    dup = 0
    for r in rows:
        no_so = s(r[0])
        if not no_so or not no_so.upper().startswith("SO"):
            continue
        if no_so in seen:
            dup += 1
            continue
        seen.add(no_so)
        out.append({
            "no_so": no_so,
            "brand": s(r[1]) or "-",
            "style": s(r[2]) or "-",
            "sumber": "IMPORT_EXCEL",
        })
    report.append("")
    report.append(f"### katalog_so — {len(out)} SO")
    report.append(f"- Duplikat NO SO dilewati: {dup}")
    return out


# --------------------------------------------------------------------------- #
# Writer CSV + SQL
# --------------------------------------------------------------------------- #
def write_csv(path: Path, rows: list[dict]):
    if not rows:
        return
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)


def write_sql(path: Path, bahan, mutasi, katalog):
    L = []
    ap = L.append
    ap("-- =====================================================================")
    ap("-- IMPOR DATA BAHAN 2026  (generated by scripts/import_excel_bahan.py)")
    ap(f"-- Dibuat: {datetime.now().isoformat(timespec='seconds')}")
    ap("-- Target: PostgreSQL / Supabase SQL Editor")
    ap("-- Idempotent: bahan_baku & katalog_so UPSERT; log_mutasi_bahan replace-by-tag")
    ap("-- PRASYARAT: tabel bahan_baku & log_mutasi_bahan sudah ada (jalankan backend")
    ap("--           sekali agar Base.metadata.create_all membuatnya). katalog_so")
    ap("--           dibuat otomatis oleh skrip ini.")
    ap("-- =====================================================================")
    ap("BEGIN;")
    ap("")
    ap("-- --- Tabel baru: katalog_so ------------------------------------------")
    ap("""CREATE TABLE IF NOT EXISTS katalog_so (
    no_so      VARCHAR(50)  PRIMARY KEY,
    brand      VARCHAR(150) NOT NULL,
    style      VARCHAR(255) NOT NULL,
    sumber     VARCHAR(50)  DEFAULT 'IMPORT_EXCEL',
    created_at TIMESTAMP    DEFAULT CURRENT_TIMESTAMP
);""")
    ap("")

    ap(f"-- --- 1. bahan_baku ({len(bahan)} item) ------------------------------")
    cols = ["id", "kode_sku", "nama_item", "kategori", "satuan", "stok_saat_ini",
            "stok_minimum", "harga_per_satuan", "lokasi_gudang", "supplier_utama",
            "warna_kode", "no_faktur_po", "tanggal_masuk", "tipe_pembayaran",
            "status_pembayaran", "jatuh_tempo"]
    for b in bahan:
        vals = [
            sqlstr(b["id"]), sqlstr(b["kode_sku"]), sqlstr(b["nama_item"]),
            sqlstr(b["kategori"]), sqlstr(b["satuan"]), sqlnum(b["stok_saat_ini"]),
            sqlnum(b["stok_minimum"]), sqlnum(b["harga_per_satuan"]),
            sqlstr(b["lokasi_gudang"]), sqlstr(b["supplier_utama"]),
            sqlstr(b["warna_kode"]), sqlstr(b["no_faktur_po"]),
            sqlstr(b["tanggal_masuk"]), sqlstr(b["tipe_pembayaran"]),
            sqlstr(b["status_pembayaran"]), sqlstr(b["jatuh_tempo"]),
        ]
        ap(f"INSERT INTO bahan_baku ({', '.join(cols)}) VALUES ({', '.join(vals)})")
        ap("  ON CONFLICT (id) DO UPDATE SET "
           "nama_item=EXCLUDED.nama_item, kategori=EXCLUDED.kategori, "
           "satuan=EXCLUDED.satuan, stok_saat_ini=EXCLUDED.stok_saat_ini, "
           "harga_per_satuan=EXCLUDED.harga_per_satuan, "
           "supplier_utama=EXCLUDED.supplier_utama, "
           "tanggal_masuk=EXCLUDED.tanggal_masuk;")
    ap("")

    ap(f"-- --- 2. katalog_so ({len(katalog)} SO) -----------------------------")
    for k in katalog:
        ap("INSERT INTO katalog_so (no_so, brand, style, sumber) VALUES ("
           f"{sqlstr(k['no_so'])}, {sqlstr(k['brand'])}, {sqlstr(k['style'])}, {sqlstr(k['sumber'])})")
        ap("  ON CONFLICT (no_so) DO UPDATE SET "
           "brand=EXCLUDED.brand, style=EXCLUDED.style;")
    ap("")

    ap(f"-- --- 3. log_mutasi_bahan ({len(mutasi)} baris) --------------------")
    ap(f"DELETE FROM log_mutasi_bahan WHERE petugas = {sqlstr(PETUGAS_TAG)};")
    mcols = ["bahan_id", "tanggal", "tipe", "jumlah", "stok_sebelum", "stok_sesudah",
             "referensi_po_spk", "catatan", "petugas"]
    for m in mutasi:
        vals = [
            sqlstr(m["bahan_id"]), sqlstr(m["tanggal"]), sqlstr(m["tipe"]),
            sqlnum(m["jumlah"]), sqlnum(m["stok_sebelum"]), sqlnum(m["stok_sesudah"]),
            sqlstr(m["referensi_po_spk"]), sqlstr(m["catatan"]), sqlstr(m["petugas"]),
        ]
        ap(f"INSERT INTO log_mutasi_bahan ({', '.join(mcols)}) VALUES ({', '.join(vals)});")
    ap("")
    ap("COMMIT;")
    ap("")
    path.write_text("\n".join(L), encoding="utf-8")


# --------------------------------------------------------------------------- #
def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    src = Path(sys.argv[1])
    outdir = Path(sys.argv[2]) if len(sys.argv) > 2 else \
        Path(__file__).resolve().parent.parent / "database" / "import_bahan_2026"
    outdir.mkdir(parents=True, exist_ok=True)

    print(f"Baca  : {src}")
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)

    report = [f"# Laporan Pembersihan Impor Bahan 2026",
              f"", f"Sumber: `{src.name}`  ",
              f"Dibuat: {datetime.now().isoformat(timespec='seconds')}", ""]

    bahan = parse_bahan(wb, report)
    valid_codes = {b["id"] for b in bahan}
    mutasi = parse_mutasi(wb, valid_codes, report)
    katalog = parse_katalog(wb, report)

    # rekonsiliasi: saldo mutasi vs stok master
    saldo = {}
    for m in mutasi:
        saldo[m["bahan_id"]] = m["stok_sesudah"]
    beda = []
    for b in bahan:
        sm = saldo.get(b["id"])
        if sm is not None and abs(sm - b["stok_saat_ini"]) > 1.0:
            beda.append((b["id"], b["stok_saat_ini"], sm))
    report.append("")
    report.append("### Rekonsiliasi stok master vs saldo mutasi")
    if beda:
        report.append(f"{len(beda)} kode selisih > 1 unit (master dipakai sbagai kebenaran; "
                      "mutasi hanya histori):")
        report.append("")
        report.append("| Kode | Stok master | Saldo dari mutasi |")
        report.append("|---|---|---|")
        for c, a, bb in beda[:40]:
            report.append(f"| {c} | {a} | {bb} |")
        if len(beda) > 40:
            report.append(f"| ... | ({len(beda)-40} lagi) | |")
    else:
        report.append("Semua cocok (selisih <= 1 unit).")

    write_csv(outdir / "01_bahan_baku.csv", bahan)
    write_csv(outdir / "02_log_mutasi_bahan.csv", mutasi)
    write_csv(outdir / "03_katalog_so.csv", katalog)
    write_sql(outdir / "import_all.sql", bahan, mutasi, katalog)
    (outdir / "LAPORAN_PEMBERSIHAN.md").write_text("\n".join(report), encoding="utf-8")

    print(f"Tulis : {outdir}")
    for p in sorted(outdir.iterdir()):
        print(f"  - {p.name}  ({p.stat().st_size:,} B)")
    print(f"\nRingkas: {len(bahan)} bahan_baku | {len(mutasi)} log_mutasi | {len(katalog)} katalog_so")


if __name__ == "__main__":
    main()
